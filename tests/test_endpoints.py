"""
Flask endpoint tests — no browser, no running server needed.
Uses Flask's built-in test client.

Upload tests send real January 2023 files so we exercise the full pipeline.
"""
import io
import os
import pytest
import openpyxl


# Paths to real files for upload tests
import sys, os
sys.path.insert(0, os.path.dirname(__file__))
from conftest import JAN_PLAT, JAN_PSP_DIR

_TRUST_PAYMENTS = os.path.join(JAN_PSP_DIR, 'TrustPayments.csv')
_EQUITY_FILE    = os.path.join(JAN_PSP_DIR, os.pardir, 'platform', 'Client Balance check.xlsx')
_DW_REPORT      = os.path.join(JAN_PSP_DIR, os.pardir, 'Deposit and Withdrawal Report.csv')


def _open(path):
    return (open(path, 'rb'), os.path.basename(path))


# ---------------------------------------------------------------------------
# Basic routing
# ---------------------------------------------------------------------------

class TestRoutes:

    def test_index_returns_200(self, flask_client):
        res = flask_client.get('/')
        assert res.status_code == 200

    def test_index_returns_html(self, flask_client):
        res = flask_client.get('/')
        assert b'MRS' in res.data

    def test_unknown_route_returns_404(self, flask_client):
        res = flask_client.get('/does-not-exist')
        assert res.status_code == 404


# ---------------------------------------------------------------------------
# /api/rates
# ---------------------------------------------------------------------------

class TestFxRates:

    def test_rates_status_key_present(self, flask_client):
        res = flask_client.get('/api/rates')
        # May fail if network is unavailable — only check the response shape
        data = res.get_json()
        assert 'status' in data

    def test_rates_returns_json(self, flask_client):
        res = flask_client.get('/api/rates')
        assert res.content_type == 'application/json'

    def test_crypto_rates_returns_json(self, flask_client):
        res = flask_client.get('/api/rates/crypto')
        assert res.content_type == 'application/json'


# ---------------------------------------------------------------------------
# /api/upload
# ---------------------------------------------------------------------------

class TestUpload:

    @pytest.fixture(autouse=True)
    def upload_result(self, flask_client):
        """POST real files once; all tests in this class reuse the response."""
        with open(JAN_PLAT, 'rb') as plat, \
             open(_TRUST_PAYMENTS, 'rb') as bank, \
             open(_EQUITY_FILE, 'rb') as eq, \
             open(_DW_REPORT, 'rb') as txn:

            res = flask_client.post('/api/upload', data={
                'bankFile':         (bank, 'TrustPayments.csv'),
                'platformFile':     (plat, 'CRM Transactions Additional info.xlsx'),
                'equityFile':       (eq,   'Client Balance check.xlsx'),
                'transactionsFile': (txn,  'Deposit and Withdrawal Report.csv'),
            }, content_type='multipart/form-data')

        self.res = res
        self.data = res.get_json()

    def test_returns_200(self):
        assert self.res.status_code == 200

    def test_status_success(self):
        assert self.data['status'] == 'success'

    def test_sources_present(self):
        assert 'sources' in self.data

    def test_bank_source_has_columns(self):
        bank = self.data['sources'].get('Bank/PSP Statements', {})
        assert len(bank.get('columns', [])) > 0

    def test_platform_source_has_columns(self):
        plat = self.data['sources'].get('Platform (CRM/MT4)', {})
        assert len(plat.get('columns', [])) > 0

    def test_message_mentions_files(self):
        assert 'files' in self.data.get('message', '').lower()

    def test_missing_bank_file_returns_400(self, flask_client):
        with open(JAN_PLAT, 'rb') as plat, \
             open(_EQUITY_FILE, 'rb') as eq, \
             open(_DW_REPORT, 'rb') as txn:
            res = flask_client.post('/api/upload', data={
                'platformFile':     (plat, 'crm.xlsx'),
                'equityFile':       (eq,   'eq.xlsx'),
                'transactionsFile': (txn,  'txn.csv'),
            }, content_type='multipart/form-data')
        assert res.status_code == 400


# ---------------------------------------------------------------------------
# /api/reconcile  (requires files uploaded first)
# ---------------------------------------------------------------------------

@pytest.fixture(scope='module')
def reconcile_result(flask_client):
    """Upload files then reconcile — reused across all reconcile/download tests."""
    with open(JAN_PLAT, 'rb') as plat, \
         open(_TRUST_PAYMENTS, 'rb') as bank, \
         open(_EQUITY_FILE, 'rb') as eq, \
         open(_DW_REPORT, 'rb') as txn:

        flask_client.post('/api/upload', data={
            'bankFile':         (bank, 'TrustPayments.csv'),
            'platformFile':     (plat, 'CRM Transactions Additional info.xlsx'),
            'equityFile':       (eq,   'Client Balance check.xlsx'),
            'transactionsFile': (txn,  'Deposit and Withdrawal Report.csv'),
        }, content_type='multipart/form-data')

    return flask_client.post('/api/reconcile')


class TestReconcile:

    def test_returns_200(self, reconcile_result):
        assert reconcile_result.status_code == 200

    def test_status_success(self, reconcile_result):
        data = reconcile_result.get_json()
        assert data['status'] == 'success', f"Reconcile error: {data}"

    def test_summary_has_required_keys(self, reconcile_result):
        summary = reconcile_result.get_json()['summary']
        required = {'total_crm_rows', 'total_bank_rows', 'total_matched',
                    'crm_unmatched', 'bank_unmatched', 'unrecon_fees', 'join_keys_used'}
        assert required.issubset(summary.keys())

    def test_crm_row_count(self, reconcile_result):
        # CRM Additional Info has 3 550 rows for Jan 2023 (not 10 255 — that's the
        # full lifecycle output which also draws from D&W Report and other sources)
        summary = reconcile_result.get_json()['summary']
        assert summary['total_crm_rows'] == 3_550

    def test_matched_plus_unmatched_equals_total(self, reconcile_result):
        s = reconcile_result.get_json()['summary']
        assert s['total_matched'] + s['crm_unmatched'] + s['bank_unmatched'] == \
               s['total_crm_rows'] + s['total_bank_rows'] - s['total_matched'], \
               "matched + crm_unmatched + bank_unmatched should account for all rows"

    def test_join_keys_identified(self, reconcile_result):
        join_info = reconcile_result.get_json()['summary']['join_keys_used']
        assert 'FAILED' not in join_info, f"Join key detection failed: {join_info}"

    def test_join_key_names_correct(self, reconcile_result):
        join_info = reconcile_result.get_json()['summary']['join_keys_used']
        assert 'psp_transaction_id' in join_info.lower()


# ---------------------------------------------------------------------------
# /api/download/lifecycle
# ---------------------------------------------------------------------------

class TestDownloadLifecycle:

    @pytest.fixture(autouse=True)
    def dl(self, flask_client, reconcile_result):  # noqa: unused — fixture dep ensures state exists
        self.res = flask_client.get('/api/download/lifecycle')

    def test_returns_200(self):
        assert self.res.status_code == 200

    def test_content_type_is_xlsx(self):
        ct = self.res.content_type
        assert 'spreadsheetml' in ct or 'officedocument' in ct, \
            f"Unexpected content type: {ct}"

    def test_file_is_valid_xlsx(self):
        wb = openpyxl.load_workbook(io.BytesIO(self.res.data))
        assert 'MT4-Transactions' in wb.sheetnames

    def test_has_expected_tabs(self):
        wb = openpyxl.load_workbook(io.BytesIO(self.res.data))
        expected = {'MT4-Transactions', 'MT4 CCY Life Cycle', 'MT4 USD per acc Life Cycle',
                    'PM-Transactions', 'PM USD Life Cycle', 'PM CCY Life Cycle', 'Mapping Rules'}
        assert expected.issubset(set(wb.sheetnames))

    def test_mt4_has_44_columns(self):
        from server import MT4_TRX_COLUMNS
        wb = openpyxl.load_workbook(io.BytesIO(self.res.data))
        ws = wb['MT4-Transactions']
        header = [cell.value for cell in ws[1]]
        assert len(header) == 44, f"Expected 44 columns, got {len(header)}"
        assert header == MT4_TRX_COLUMNS

    def test_pm_has_31_columns(self):
        from server import PM_TRX_COLUMNS
        wb = openpyxl.load_workbook(io.BytesIO(self.res.data))
        ws = wb['PM-Transactions']
        header = [cell.value for cell in ws[1]]
        assert len(header) == 31, f"Expected 31 columns, got {len(header)}"
        assert header == PM_TRX_COLUMNS

    def test_ccy_lifecycle_has_4_columns(self):
        wb = openpyxl.load_workbook(io.BytesIO(self.res.data))
        ws = wb['MT4 CCY Life Cycle']
        header = [cell.value for cell in ws[1]]
        assert header == ['Client Account', 'Currency', 'Attribute', 'Amount']

    def test_has_data_rows(self):
        wb = openpyxl.load_workbook(io.BytesIO(self.res.data))
        ws = wb['MT4-Transactions']
        assert ws.max_row > 1, "MT4-Transactions sheet has no data rows"


# ---------------------------------------------------------------------------
# /api/download/balances
# ---------------------------------------------------------------------------

class TestDownloadBalances:

    @pytest.fixture(autouse=True)
    def dl(self, flask_client, reconcile_result):  # noqa: unused — fixture dep ensures state exists
        self.res = flask_client.get('/api/download/balances')

    def test_returns_200(self):
        assert self.res.status_code == 200

    def test_content_type_is_xlsx(self):
        ct = self.res.content_type
        assert 'spreadsheetml' in ct or 'officedocument' in ct

    def test_file_is_valid_xlsx(self):
        wb = openpyxl.load_workbook(io.BytesIO(self.res.data))
        assert 'Balances' in wb.sheetnames

    def test_has_correct_columns(self):
        wb = openpyxl.load_workbook(io.BytesIO(self.res.data))
        ws = wb['Balances']
        header = [cell.value for cell in ws[1]]
        assert header == ['Currency', 'Equity', 'Equity EUR', 'Equity USD', 'Perc']

    def test_has_data_rows(self):
        wb = openpyxl.load_workbook(io.BytesIO(self.res.data))
        ws = wb['Balances']
        assert ws.max_row > 1


# ---------------------------------------------------------------------------
# /api/download/* without prior reconcile
# ---------------------------------------------------------------------------

class TestDownloadWithoutReconcile:

    def test_lifecycle_without_state_returns_400(self, flask_client, tmp_path):
        """If _recon_state.pkl doesn't exist the endpoint should return 400."""
        import server
        original = server.STATE_FILE
        server.STATE_FILE = str(tmp_path / 'nonexistent.pkl')
        try:
            res = flask_client.get('/api/download/lifecycle')
            assert res.status_code == 400
        finally:
            server.STATE_FILE = original

    def test_balances_without_state_returns_400(self, flask_client, tmp_path):
        import server
        original = server.STATE_FILE
        server.STATE_FILE = str(tmp_path / 'nonexistent.pkl')
        try:
            res = flask_client.get('/api/download/balances')
            assert res.status_code == 400
        finally:
            server.STATE_FILE = original
