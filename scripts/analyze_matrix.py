"""
Script 3: Analyze the PSPs - AGREEMENT MATRIX.xlsx at the root of AGREEMENTS folder.
Output: scripts/output/matrix_analysis.txt
Run: python scripts/analyze_matrix.py
"""
from pathlib import Path
import openpyxl

MATRIX_FILE = Path(__file__).parent.parent / "relevant-data" / "AGREEMENTS" / "PSPs - AGREEMENT MATRIX.xlsx"
OUTPUT_FILE = Path(__file__).parent / "output" / "matrix_analysis.txt"
MAX_ROWS    = 30   # rows to print per sheet


def cell_val(cell):
    v = cell.value
    if v is None:
        return ""
    return str(v).strip().replace("\n", " ")


def main():
    if not MATRIX_FILE.exists():
        print(f"ERROR: File not found: {MATRIX_FILE}")
        return

    wb = openpyxl.load_workbook(MATRIX_FILE, read_only=True, data_only=True)
    lines = []

    lines.append(f"MATRIX FILE: {MATRIX_FILE.name}")
    lines.append(f"Sheets: {wb.sheetnames}\n")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append("=" * 70)
        lines.append(f"SHEET: {sheet_name}")
        lines.append("=" * 70)

        rows = list(ws.iter_rows(max_row=MAX_ROWS + 1))
        if not rows:
            lines.append("  (empty sheet)")
            continue

        # Header row
        headers = [cell_val(c) for c in rows[0]]
        lines.append(f"Columns ({len(headers)}): {headers}\n")

        # Data rows
        for i, row in enumerate(rows[1:MAX_ROWS + 1], start=1):
            vals = [cell_val(c) for c in row]
            if not any(vals):
                continue
            lines.append(f"  Row {i:>3}: {vals}")

        total_rows = ws.max_row or "?"
        lines.append(f"\n  (Sheet has ~{total_rows} rows total, showing first {MAX_ROWS})")

    output = "\n".join(lines)
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        f.write(output)

    print(output.encode("ascii", errors="replace").decode("ascii"))
    print(f"\n\nFull analysis written to: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
