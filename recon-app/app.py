import hmac
import io
import os
import threading
import uuid
import functools
from datetime import date, datetime
from flask import Flask, render_template, request, jsonify, send_file, abort, Response, redirect, url_for
from dotenv import load_dotenv

load_dotenv()

# -- Live MT5 push store (updated by POST /cro/feed from Windows bridge) ------
_CRO_LIVE: dict = {}
_CRO_LIVE_LOCK = threading.Lock()
_CRO_LIVE_MAX_AGE_S = 90  # after 90s without a push, fall back to Dealio snapshot

import queries
import ai_parse
import openpyxl

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET", "dev")

@app.context_processor
def inject_request():
    return {"request": request}

try:
    queries.ensure_fee_tables()
except Exception as e:
    print(f"WARNING: Could not create fee tables: {e}")

# Pre-warm the "all" span cache in the background after startup.
# Warm 1Y and All caches 10s after start (both use background computation path).
# Users hitting these spans immediately get "computing" overlay + auto-retry.
def _warm_wide_span_caches():
    import time, datetime as _dt
    time.sleep(10)
    today = _dt.date.today()
    spans = [
        (today - _dt.timedelta(days=365), today + _dt.timedelta(days=1), "1y"),
        (today - _dt.timedelta(days=730), today + _dt.timedelta(days=1), "2y"),
        (_dt.date(2021, 1, 1),           today + _dt.timedelta(days=1), "all"),
    ]
    for df, dt, label in spans:
        try:
            queries.client_list(df, dt)
            print(f"[warmup] client_list({label}) populated")
        except Exception as e:
            print(f"[warmup] client_list({label}) failed: {e}")

import threading as _threading
_threading.Thread(target=_warm_wide_span_caches, daemon=True).start()


# NOTE: CRO dashboard is now served entirely from live push data (cro-bridge
# container), so no Dealio-based warmup is needed.

_RECON_USER  = os.environ.get("RECON_USER",  "")
_RECON_PASS  = os.environ.get("RECON_PASS",  "")
_FEES_USER   = os.environ.get("FEES_USER",   "")
_FEES_PASS   = os.environ.get("FEES_PASS",   "")
_FX_USER     = os.environ.get("FX_USER",     "")
_FX_PASS     = os.environ.get("FX_PASS",     "")
_ADMIN_USER      = os.environ.get("ADMIN_USER",      "")
_ADMIN_PASS      = os.environ.get("ADMIN_PASS",      "")
_RETENTION_USER  = os.environ.get("RETENTION_USER",  "")
_RETENTION_PASS  = os.environ.get("RETENTION_PASS",  "")
_CRO_USER        = os.environ.get("CRO_USER",        "")
_CRO_PASS        = os.environ.get("CRO_PASS",        "")


def _unauthorized(realm):
    return Response(
        "Authentication required.",
        401,
        {"WWW-Authenticate": f'Basic realm="{realm}"'},
    )


def require_recon_auth(f):
    @functools.wraps(f)
    def wrapper(*args, **kwargs):
        if not _RECON_USER or not _RECON_PASS:
            abort(500, "RECON_USER and RECON_PASS env vars not set.")
        auth = request.authorization
        if not auth or auth.username != _RECON_USER or auth.password != _RECON_PASS:
            return _unauthorized("CMT Reconciliation")
        return f(*args, **kwargs)
    return wrapper


def require_fees_auth(f):
    @functools.wraps(f)
    def wrapper(*args, **kwargs):
        if not _FEES_USER or not _FEES_PASS:
            abort(500, "FEES_USER and FEES_PASS env vars not set.")
        auth = request.authorization
        if not auth or auth.username != _FEES_USER or auth.password != _FEES_PASS:
            return _unauthorized("CMT Fee Processor")
        return f(*args, **kwargs)
    return wrapper


def require_fx_auth(f):
    @functools.wraps(f)
    def wrapper(*args, **kwargs):
        if not _FX_USER or not _FX_PASS:
            abort(500, "FX_USER and FX_PASS env vars not set.")
        auth = request.authorization
        if not auth or auth.username != _FX_USER or auth.password != _FX_PASS:
            return _unauthorized("CMT FX Rates")
        return f(*args, **kwargs)
    return wrapper


def require_admin_auth(f):
    @functools.wraps(f)
    def wrapper(*args, **kwargs):
        if not _ADMIN_USER or not _ADMIN_PASS:
            abort(500, "ADMIN_USER and ADMIN_PASS env vars not set.")
        auth = request.authorization
        if not auth or auth.username != _ADMIN_USER or auth.password != _ADMIN_PASS:
            return _unauthorized("CMT Admin")
        return f(*args, **kwargs)
    return wrapper


def require_retention_auth(f):
    @functools.wraps(f)
    def wrapper(*args, **kwargs):
        if not _RETENTION_USER or not _RETENTION_PASS:
            abort(500, "RETENTION_USER and RETENTION_PASS env vars not set.")
        auth = request.authorization
        if not auth or auth.username != _RETENTION_USER or auth.password != _RETENTION_PASS:
            return _unauthorized("CMT Retention")
        return f(*args, **kwargs)
    return wrapper


def require_cro_auth(f):
    @functools.wraps(f)
    def wrapper(*args, **kwargs):
        if not _CRO_USER or not _CRO_PASS:
            abort(500, "CRO_USER and CRO_PASS env vars not set.")
        auth = request.authorization
        if not auth or auth.username != _CRO_USER or auth.password != _CRO_PASS:
            return _unauthorized("CMT CRO")
        return f(*args, **kwargs)
    return wrapper


@app.route("/")
@require_recon_auth
def index():
    try:
        months = queries.available_months()
    except Exception as e:
        months = []
    return render_template("index.html", months=months)


@app.route("/recon/<month>")
@require_recon_auth
def recon(month):
    """Returns page shell immediately; data is loaded via /recon/<month>/groups."""
    try:
        int(month[:4]); int(month[5:7])
    except (ValueError, IndexError):
        abort(400)
    try:
        months = queries.available_months()
    except Exception as e:
        return render_template("index.html", months=[], error=str(e))
    status_filter = request.args.get("status", "all")
    hide_noncash  = request.args.get("hide_noncash") == "1"
    return render_template("recon.html", month=month, months=months,
                           status_filter=status_filter, hide_noncash=hide_noncash,
                           rows=[], groups=[], stats=None, cache_age=None)


@app.route("/recon/<month>/groups")
@require_recon_auth
def recon_groups(month):
    """Heavy JSON endpoint — called by JS after page shell has loaded."""
    try:
        year, mon = int(month[:4]), int(month[5:7])
    except (ValueError, IndexError):
        abort(400)

    hide_noncash  = request.args.get("hide_noncash") == "1"
    status_filter = request.args.get("status", "all")

    try:
        rows      = queries.reconcile(year, mon)
        cache_age = queries.cache_age(year, mon)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    if hide_noncash:
        rows = [r for r in rows if not (
            r["crm_cash_dep"] == 0 and r["crm_cash_with"] == 0
            and (r["crm_noncash_in"] or r["crm_noncash_out"])
        )]

    stats = queries.summary_stats(rows)

    try:
        groups = queries.reconcile_grouped(year, mon)
    except Exception:
        groups = []

    if hide_noncash:
        groups = [{**g, "logins": [r for r in g["logins"] if not (
            r["crm_cash_dep"] == 0 and r["crm_cash_with"] == 0
            and (r["crm_noncash_in"] or r["crm_noncash_out"])
        )]} for g in groups]
        groups = [g for g in groups if g["logins"]]

    if status_filter != "all":
        groups = [g for g in groups if g["agg"]["status"] == status_filter]

    return jsonify({"groups": groups, "stats": stats, "cache_age": cache_age})


@app.route("/clients/equity-report")
@require_recon_auth
def clients_equity_report():
    import datetime as _dt
    span = request.args.get("span", "1y")
    today = _dt.date.today()
    # Custom explicit date range overrides span
    df_str = request.args.get("date_from", "")
    dt_str = request.args.get("date_to",   "")
    if df_str and dt_str:
        try:
            date_from = _dt.date.fromisoformat(df_str)
            date_to   = _dt.date.fromisoformat(dt_str) + _dt.timedelta(days=1)
            span = f"{df_str}_to_{dt_str}"
        except ValueError:
            date_from = today - _dt.timedelta(days=365)
            date_to   = today + _dt.timedelta(days=1)
    else:
        span_map = {"1w":7,"1m":31,"3m":92,"6m":183,"1y":365,"2y":730,"all":0}
        if span == "all":
            date_from = _dt.date(2021,1,1)
        else:
            date_from = today - _dt.timedelta(days=span_map.get(span,365))
        date_to = today + _dt.timedelta(days=1)
    try:
        rows = queries.equity_report(date_from, date_to)
    except Exception as e:
        abort(500, str(e))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Equity Report"
    ws.append(["Login","Name","CID","Currency",
               "Open Balance","Open Equity",
               "Deposits (USD)","Withdrawals (USD)","Realised P&L (USD)",
               "Close Balance","Close Equity","Last Active"])
    for r in rows:
        ws.append([r["login"],r["name"] or "",r["cid"] or "",r["currency"],
                   r["open_balance"],r["open_equity"],
                   r["deposits"],r["withdrawals"],r["realised_pnl"],
                   r["close_balance"],r["close_equity"],r["last_active"]])
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(w+4, 40)
    buf = __import__("io").BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"equity_report_{span}_{date_to}.xlsx")


@app.route("/clients/data")
@require_recon_auth
def clients_data():
    """JSON endpoint for client list — used by the JS virtual table."""
    import datetime as _dt
    span  = request.args.get("span", "1y")
    today = _dt.date.today()
    # Custom explicit date range overrides span
    df_str = request.args.get("date_from", "")
    dt_str = request.args.get("date_to",   "")
    if df_str and dt_str:
        try:
            date_from = _dt.date.fromisoformat(df_str)
            date_to   = _dt.date.fromisoformat(dt_str) + _dt.timedelta(days=1)
        except ValueError:
            date_from = today - _dt.timedelta(days=365)
            date_to   = today + _dt.timedelta(days=1)
    else:
        span_map = {"1w":7,"1m":31,"3m":92,"6m":183,"1y":365,"2y":730,"all":0}
        if span == "all":
            date_from = _dt.date(2021,1,1)
        else:
            date_from = today - _dt.timedelta(days=span_map.get(span,365))
        date_to = today + _dt.timedelta(days=1)
    error_msg = None
    try:
        rows = queries.client_list(date_from, date_to)
        cache_age = queries.cache_age_key(f"client_list:{date_from}:{date_to}")
    except Exception as e:
        rows = []
        cache_age = None
        error_msg = str(e)
        import traceback, sys
        print(f"[clients/data] ERROR {span}/{date_from}→{date_to}: {e}", file=sys.stderr)
        traceback.print_exc(file=sys.stderr)
    # If empty result AND background computation is running, tell frontend to retry
    computing = not rows and queries.is_client_list_computing()
    resp = {"rows": rows, "cache_age": cache_age,
            "date_from": str(date_from), "date_to": str(date_to),
            "computing": computing, "error": error_msg}
    if computing:
        prog = queries.get_client_list_progress(date_from, date_to)
        resp["stage"]     = prog["stage"] if prog else "Starting computation\u2026"
        resp["stage_num"] = prog["num"]   if prog else 0
    return jsonify(resp)


@app.route("/clients")
@require_recon_auth
def clients():
    span = request.args.get("span", "1y")
    span_labels = {"1w":"1 Week","1m":"1 Month","3m":"3 Months","6m":"6 Months","1y":"1 Year","2y":"2 Years","all":"All Time"}
    return render_template("clients.html", span=span,
                           span_label=span_labels.get(span, span))


@app.route("/cid/<cid>/floating")
@require_recon_auth
def cid_floating(cid):
    """Live floating P&L for a CID — polled by JS every 30s."""
    account_map, _ = queries._load_praxis_account_map()
    logins = account_map.get(str(cid).strip(), [])
    if not logins:
        return jsonify({})
    try:
        from db import dealio
        with dealio() as cur:
            cur.execute("SET statement_timeout = 10000")
            ph = ",".join(["%s"] * len(logins))
            cur.execute(f"""
                SELECT DISTINCT ON (login) login,
                    convertedfloatingpnl AS floating,
                    balance, equity, date
                FROM dealio.daily_profits
                WHERE login IN ({ph})
                ORDER BY login, date DESC
            """, logins)
            result = {}
            for r in cur.fetchall():
                result[str(r["login"])] = {
                    "floating": round(float(r["floating"] or 0), 2),
                    "balance":  round(float(r["balance"] or 0), 2),
                    "equity":   round(float(r["equity"] or 0), 2),
                    "date":     str(r["date"]) if r["date"] else "",
                }
            return jsonify(result)
    except Exception:
        return jsonify({})


@app.route("/cid/<cid>")
@require_recon_auth
def cid_detail(cid):
    import datetime as _dt
    span = request.args.get("span", "1m")
    ref  = request.args.get("ref", "")

    today = _dt.date.today()
    last = request.args.get("last", "")   # YYYY-MM anchor from client list

    # range_end: use ref month end if coming from recon page, else today+1
    if ref:
        try:
            ry, rm = int(ref[:4]), int(ref[5:7])
            range_end = _dt.date(ry+1,1,1) if rm==12 else _dt.date(ry,rm+1,1)
        except Exception:
            range_end = today + _dt.timedelta(days=1)
    else:
        range_end = today + _dt.timedelta(days=1)

    span_map = {"1m":31,"3m":92,"6m":183,"1y":365}
    span_labels = {"1m":"1 Month","3m":"3 Months","6m":"6 Months","1y":"1 Year","all":"All Time"}

    if span == "all":
        range_start = _dt.date(2021, 1, 1)
    else:
        d = range_end - _dt.timedelta(days=span_map.get(span, 31))
        range_start = _dt.date(d.year, d.month, 1)

    try:
        profile = queries.cid_full_profile(cid, range_start, range_end)
    except Exception:
        profile = {"cid": cid, "name":"—","email":"—","mt4_accounts":[],
                   "praxis_txs":[],"crm_by_login":{},"mt4_by_login":{},
                   "summary":{"praxis_deposits":0,"praxis_withdrawals":0,
                               "crm_cash_dep":0,"crm_cash_with":0,"diff":0}}

    return render_template("cid_detail.html",
        cid=cid, profile=profile, span=span,
        span_label=span_labels.get(span, span),
        range_start=str(range_start), range_end=str(range_end), ref=ref)


@app.route("/client/<int:login>")
@require_recon_auth
def client_detail(login):
    import datetime as _dt
    span = request.args.get("span", "1m")
    ref  = request.args.get("ref", "")   # optional reference month e.g. "2026-04"

    # Determine date range
    today = _dt.date.today()
    if ref:
        try:
            ry, rm = int(ref[:4]), int(ref[5:7])
            range_end = (_dt.date(ry+1,1,1) if rm==12 else _dt.date(ry,rm+1,1))
        except Exception:
            range_end = _dt.date(today.year, today.month, 1) + _dt.timedelta(days=32)
            range_end = _dt.date(range_end.year, range_end.month, 1)
    else:
        range_end = _dt.date(today.year, today.month, 1) + _dt.timedelta(days=32)
        range_end = _dt.date(range_end.year, range_end.month, 1)

    span_map = {
        "1m":  _dt.timedelta(days=31),
        "3m":  _dt.timedelta(days=92),
        "6m":  _dt.timedelta(days=183),
        "1y":  _dt.timedelta(days=365),
    }
    if span == "all":
        range_start = _dt.date(2021, 1, 1)
    else:
        d = range_end - span_map.get(span, _dt.timedelta(days=31))
        range_start = _dt.date(d.year, d.month, 1)

    span_labels = {"1m":"1 Month","3m":"3 Months","6m":"6 Months","1y":"1 Year","all":"All Time"}

    try:
        crm_rows    = queries.client_crm_detail(login, range_start, range_end)
        mt4_rows    = queries.client_mt4_detail(login, range_start, range_end)
        praxis_rows = queries.client_praxis_detail(login, range_start, range_end)
    except Exception as e:
        crm_rows = mt4_rows = praxis_rows = []

    # Summary totals
    crm_cash   = sum(r["total_usd"] for r in crm_rows
                     if r["is_cash"] and r.get("transactionapproval") == "Approved"
                     and r.get("transactiontype") in ("Deposit","TransferIn"))
    crm_with   = sum(r["total_usd"] for r in crm_rows
                     if r["is_cash"] and r.get("transactionapproval") == "Approved"
                     and r.get("transactiontype") in ("Withdrawal","Withdraw","TransferOut"))
    praxis_dep = sum(r["usd_amount"] for r in praxis_rows if r["direction"] == "payment")
    praxis_with= sum(r["usd_amount"] for r in praxis_rows
                     if r["direction"] in ("withdrawal","payout"))

    try:
        return render_template("client_detail.html",
            login=login, span=span, span_label=span_labels.get(span, span),
            range_start=str(range_start), range_end=str(range_end), ref=ref,
            crm_rows=crm_rows, mt4_rows=mt4_rows, praxis_rows=praxis_rows,
            crm_cash=round(crm_cash,2), crm_with=round(crm_with,2),
            praxis_dep=round(praxis_dep,2), praxis_with=round(praxis_with,2),
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return f"<pre>Error rendering client detail for login {login}:\n{e}\n\n{traceback.format_exc()}</pre>", 500


@app.route("/recon/<month>/crm-txns/<int:login>")
@require_recon_auth
def recon_crm_txns(month, login):
    """JSON breakdown of CRM transactions for a login in a month (hover popover)."""
    try:
        year, mon = int(month[:4]), int(month[5:7])
    except (ValueError, IndexError):
        abort(400)
    rows = queries.login_crm_transactions(year, mon, login)
    return jsonify(rows)


@app.route("/recon/<month>/praxis")
@require_recon_auth
def recon_praxis(month):
    import datetime as _dt
    try:
        year, mon = int(month[:4]), int(month[5:7])
    except (ValueError, IndexError):
        abort(400)

    span = request.args.get("span", "1m")   # 1m 3m 6m 1y all

    # Compute date range from span
    month_end   = (_dt.date(year + 1, 1, 1) if mon == 12
                   else _dt.date(year, mon + 1, 1))
    span_labels = {"1m": "1 Month", "3m": "3 Months", "6m": "6 Months",
                   "1y": "1 Year",  "all": "All Time"}
    if span == "3m":
        d = month_end - _dt.timedelta(days=92)
        date_from = _dt.date(d.year, d.month, 1)
    elif span == "6m":
        d = month_end - _dt.timedelta(days=183)
        date_from = _dt.date(d.year, d.month, 1)
    elif span == "1y":
        d = month_end - _dt.timedelta(days=365)
        date_from = _dt.date(d.year, d.month, 1)
    elif span == "all":
        date_from = _dt.date(2021, 1, 1)   # Praxis data starts 2021
    else:
        date_from = _dt.date(year, mon, 1)

    try:
        tree   = queries.praxis_client_tree(year, mon,
                                            date_from=date_from, date_to=month_end)
        months = queries.available_months()
    except Exception as e:
        tree   = []
        months = []
    praxis_error = queries.get_praxis_error()
    return render_template("praxis_tree.html", month=month, tree=tree, months=months,
                           span=span, span_label=span_labels.get(span, span),
                           date_from=str(date_from), date_to=str(month_end),
                           praxis_error=praxis_error)


@app.route("/recon/<month>/refresh", methods=["POST"])
@require_recon_auth
def recon_refresh(month):
    try:
        year, mon = int(month[:4]), int(month[5:7])
    except (ValueError, IndexError):
        abort(400)
    queries.cache_invalidate(year, mon)
    return redirect(url_for("recon", month=month,
                            status=request.form.get("status", "all"),
                            hide_noncash=request.form.get("hide_noncash", "0")))


@app.route("/fx")
@require_fx_auth
def fx_rates():
    return render_template("fx.html", fx_groups=queries.FX_GROUPS)


_FX_PERIOD_MAP = {
    "5s":  1,     # nearest available tick ~ 1 min ago (5s not in history, use 1m)
    "30s": 1,
    "1m":  1,
    "5m":  5,
    "15m": 15,
    "1h":  60,
    "4h":  240,
    "1d":  1440,
    "1w":  10080,
    "1mo": 43200,
}


@app.route("/fx/api/rates")
@require_fx_auth
def fx_api_rates():
    """Returns live rates AND reference prices for the requested period in one call,
    eliminating any timing gap between the two fetches."""
    symbols    = request.args.getlist("s") or None
    ref_period = request.args.get("ref_period", "1h")
    minutes    = _FX_PERIOD_MAP.get(ref_period, 60)
    try:
        rows = queries.get_live_fx_rates(symbols)
        refs = queries.get_reference_fx_rates(minutes, symbols)
        age  = queries.cache_age_key("fx_live:" + ",".join(sorted(symbols or queries.FX_ALL_SYMBOLS)))
        return jsonify({"rates": rows, "references": refs,
                        "ref_period": ref_period, "cache_age": age})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/fx/api/ohlc/<symbol>")
@require_fx_auth
def fx_api_ohlc(symbol):
    period = request.args.get("period", "1d")
    try:
        rows = queries.get_fx_ohlc(symbol, period)
        return jsonify({"symbol": symbol, "period": period, "data": rows})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/fx/api/history/<symbol>")
@require_fx_auth
def fx_api_history(symbol):
    hours = request.args.get("hours", 168, type=int)
    try:
        rows = queries.get_fx_history(symbol, hours)
        return jsonify({"symbol": symbol, "hours": hours, "data": rows})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


def _expected_fee(usd_amount: float, payment_method: str, direction: str,
                  fee_rules_by_psp: dict, payment_processor: str) -> float:
    """
    Calculate the expected fee from psp_fee_rules for a single Praxis transaction.
    fee_rules_by_psp: {psp_name_lower: [rule, ...]} — loaded once per export run.
    Returns expected fee in USD, or 0.0 if no matching rule found.
    """
    if not usd_amount:
        return 0.0
    fee_type = "Deposit" if direction == "payment" else "Withdrawal"
    pm_norm  = (payment_method or "").lower()

    # Find matching PSP rules (fuzzy match on payment_processor)
    rules = []
    for psp_key, psp_rules in fee_rules_by_psp.items():
        if psp_key in (payment_processor or "").lower() or \
           (payment_processor or "").lower() in psp_key:
            rules = psp_rules
            break

    best_rule = None
    for rule in rules:
        if rule.get("fee_type") != fee_type:
            continue
        rule_pm = (rule.get("payment_method") or "").lower()
        if rule_pm and rule_pm not in pm_norm and pm_norm not in rule_pm:
            continue
        best_rule = rule
        break   # first match wins

    if not best_rule:
        return 0.0

    kind = best_rule.get("fee_kind", "percentage")
    pct  = float(best_rule.get("pct_rate") or 0)
    fix  = float(best_rule.get("fixed_amount") or 0)

    if kind == "percentage":
        return round(usd_amount * pct, 4)
    if kind == "fixed":
        return round(fix, 4)
    if kind == "fixed_plus_pct":
        return round(fix + usd_amount * pct, 4)
    if kind == "tiered":
        tiers = sorted(best_rule.get("tiers", []), key=lambda t: t.get("volume_from", 0))
        for tier in reversed(tiers):
            if usd_amount >= tier.get("volume_from", 0):
                return round(usd_amount * float(tier.get("pct_rate", 0)), 4)
    return 0.0


# ── Bank ↔ CRM Matching Dashboard ───────────────────────────────────────────

@app.route("/recon/<month>/bank-match")
@require_recon_auth
def bank_match_page(month):
    try:
        year, mon = int(month[:4]), int(month[5:7])
    except (ValueError, IndexError):
        abort(400)
    bank_txns = queries.bank_transactions_for_period(year, mon)
    crm_txns  = queries.crm_cash_transactions_individual(year, mon)
    months    = queries.available_months()
    matched   = [b for b in bank_txns if b.get("match_status") == "matched"]
    unmatched = [b for b in bank_txns if (b.get("match_status") or "unmatched") == "unmatched"]
    excluded  = [b for b in bank_txns if b.get("match_status") == "excluded"]
    return render_template("bank_reconcile.html",
                           month=month, months=months,
                           bank_txns=bank_txns, crm_txns=crm_txns,
                           matched=matched, unmatched=unmatched, excluded=excluded)


@app.route("/recon/<month>/bank-match/auto", methods=["POST"])
@require_recon_auth
def bank_match_auto(month):
    try:
        year, mon = int(month[:4]), int(month[5:7])
    except (ValueError, IndexError):
        abort(400)
    stats = queries.auto_match_bank_to_crm(year, mon)
    queries.cache_invalidate(year, mon)
    return redirect(f"/recon/{month}/bank-match")


@app.route("/recon/<month>/bank-match/save", methods=["POST"])
@require_recon_auth
def bank_match_save(month):
    bank_tx_id = request.form.get("bank_tx_id", type=int)
    crm_tx_id  = request.form.get("crm_tx_id", type=int)
    if not bank_tx_id or not crm_tx_id:
        abort(400)
    from db import fees_db as _fdb
    # Look up the login for this CRM transaction so bank_recon_summary works without CRM
    crm_login = None
    try:
        year_s, mon_s = int(month[:4]), int(month[5:7])
        crm_txns = queries.crm_cash_transactions_individual(year_s, mon_s)
        crm_login = next((int(c["login"]) for c in crm_txns
                          if c.get("transactionid") == crm_tx_id), None)
    except Exception:
        pass
    with _fdb() as conn:
        conn.execute("""
            UPDATE bank_transactions
            SET matched_crm_id=?, match_confidence=1.0, match_status='manual', matched_login=?
            WHERE id=?
        """, (crm_tx_id, crm_login, bank_tx_id))
    try:
        year, mon = int(month[:4]), int(month[5:7])
        queries.cache_invalidate(year, mon)
    except Exception:
        pass
    return redirect(f"/recon/{month}/bank-match")


@app.route("/recon/<month>/bank-match/exclude", methods=["POST"])
@require_recon_auth
def bank_match_exclude(month):
    bank_tx_id = request.form.get("bank_tx_id", type=int)
    if not bank_tx_id:
        abort(400)
    from db import fees_db as _fdb
    with _fdb() as conn:
        conn.execute(
            "UPDATE bank_transactions SET match_status='excluded' WHERE id=?",
            (bank_tx_id,)
        )
    return redirect(f"/recon/{month}/bank-match")


@app.route("/recon/<month>/bank-match/unmatch", methods=["POST"])
@require_recon_auth
def bank_match_unmatch(month):
    bank_tx_id = request.form.get("bank_tx_id", type=int)
    if not bank_tx_id:
        abort(400)
    from db import fees_db as _fdb
    with _fdb() as conn:
        conn.execute("""
            UPDATE bank_transactions
            SET matched_crm_id=NULL, match_confidence=NULL, match_status='unmatched'
            WHERE id=?
        """, (bank_tx_id,))
    try:
        year, mon = int(month[:4]), int(month[5:7])
        queries.cache_invalidate(year, mon)
    except Exception:
        pass
    return redirect(f"/recon/{month}/bank-match")


@app.route("/recon/<month>/export")
@require_recon_auth
def export(month):
    try:
        year, mon = int(month[:4]), int(month[5:7])
    except (ValueError, IndexError):
        abort(400)

    # ── Load all data (parallel benefit from cache) ──────────────────────
    recon_rows   = queries.reconcile(year, mon)
    equity_rows  = queries.equity_by_client(year, mon)
    crm_txs      = queries.crm_transaction_list(year, mon)
    praxis_txs   = queries.praxis_transaction_list(year, mon)
    pnl_rows     = queries.profitability_by_day(year, mon)
    psp_balances = queries.psp_balance_at_month_end(year, mon)

    # Load fee rules for expected fee calculation (keyed by psp_name lower)
    all_agreements = queries.get_all_agreements()
    fee_rules_by_psp = {}
    for agr in all_agreements:
        rules = queries.get_fee_rules(agr["id"])
        fee_rules_by_psp[agr["psp_name"].lower()] = rules

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary (fixed keys) ────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.append(["Login", "MT4 Net (USD)", "CRM Cash Net (USD)", "CRM Deposits (USD)",
                 "CRM Withdrawals (USD)", "Praxis Net (USD)", "Praxis Deposits (USD)",
                 "Praxis Withdrawals (USD)", "Bank Net", "Bank Matched Txns",
                 "Difference (MT4 vs CRM)",
                 "Status", "Payment Methods", "CRM Tx Count", "Praxis Tx Count", "Currency"])
    for r in recon_rows:
        ws1.append([
            r["login"], r["mt4_net"], r["crm_cash_net"],
            r["crm_cash_dep"], r["crm_cash_with"],
            r["praxis_net"], r["praxis_deposits"], r["praxis_withdrawals"],
            r.get("bank_net", 0), r.get("bank_matched", 0),
            r["difference"], r["status"],
            r["payment_methods"], r["tx_count"], r["praxis_tx_count"], r["currency"],
        ])

    # ── Sheet 2: Equity at Month End ─────────────────────────────────────
    ws2 = wb.create_sheet("Equity at Month End")
    ws2.append(["Login", "Currency", "Balance", "Equity", "Date"])
    for r in equity_rows:
        ws2.append([r["login"], r["currency"],
                    float(r["balance"] or 0), float(r["equity"] or 0),
                    str(r["date"])])

    # ── Sheet 3: Transactions ─────────────────────────────────────────────
    ws3 = wb.create_sheet("Transactions")
    ws3.append([
        "Login", "Date", "Source", "Type / Direction",
        "Payment Method", "PSP / Processor",
        "Amount (Local)", "Currency", "Amount (USD)",
        "Actual Fee (USD)", "Expected Fee (USD)", "Fee Variance (USD)",
        "Email", "Customer Name",
        "CRM Reference", "Praxis TID", "Order ID", "Approval / Status"
    ])
    for r in crm_txs:
        ws3.append([
            r["login"],
            str(r["confirmation_time"])[:16] if r["confirmation_time"] else "",
            "CRM", r["transactiontype"],
            r["payment_method"], "",
            float(r["usdamount"] or 0), "USD", float(r["usdamount"] or 0),
            "", "", "",
            "", "",
            r["transactionid"], "", "", r["transactionapproval"],
        ])
    for r in praxis_txs:
        usd_amt = float(r["usd_amount"] or 0)
        actual  = float(r["fee_actual"] or 0)
        proc    = r.get("payment_processor") or ""
        pm      = r.get("payment_method") or ""
        dirn    = r.get("direction") or ""
        expected = _expected_fee(usd_amt, pm, dirn, fee_rules_by_psp, proc)
        login = r.get("login")
        try:
            login = int(login) if login else None
        except (ValueError, TypeError):
            login = None
        ws3.append([
            login,
            str(r["inserted_at"])[:16] if r["inserted_at"] else "",
            "Praxis", dirn,
            pm, proc,
            float(r["amount_local"] or 0), r["currency"], usd_amt,
            round(actual, 2), round(expected, 2), round(actual - expected, 2),
            r.get("email") or "",
            f"{r.get('customer_first_name','')} {r.get('customer_last_name','')}".strip(),
            "", r.get("tid"), r.get("session_order_id"), "",
        ])

    # ── Sheet 4: Profitability by Day ────────────────────────────────────
    ws4 = wb.create_sheet("Profitability by Day")
    ws4.append(["Login", "Date", "Currency", "Realised P&L", "Unrealised P&L (EOD)",
                 "Balance", "Equity"])
    for r in pnl_rows:
        ws4.append([
            r["login"], str(r["date"]), r["currency"],
            float(r["realised_pnl"] or 0), float(r["unrealised_pnl_eod"] or 0),
            float(r["balance"] or 0), float(r["equity"] or 0),
        ])

    # ── Sheet 5: PSP Balances ─────────────────────────────────────────────
    ws5 = wb.create_sheet("PSP Balances")
    ws5.append(["PSP", "Currency", "Gross Volume (USD)", "Deposits (USD)",
                 "Withdrawals (USD)", "Net (USD)",
                 "Actual Fees (USD)", "Expected Fees (USD)", "Fee Variance (USD)",
                 "Transaction Count"])
    psp_expected: dict = {}
    for r in praxis_txs:
        proc = (r.get("payment_processor") or "unknown")
        usd_amt = float(r["usd_amount"] or 0)
        dirn    = r.get("direction") or ""
        pm      = r.get("payment_method") or ""
        exp     = _expected_fee(usd_amt, pm, dirn, fee_rules_by_psp, proc)
        psp_expected[proc] = psp_expected.get(proc, 0) + exp

    for r in psp_balances:
        psp     = r["psp"] or "unknown"
        actual  = float(r["actual_fees_usd"] or 0)
        expected = round(psp_expected.get(psp, 0), 2)
        ws5.append([
            psp, r["currency"],
            float(r["gross_volume_usd"] or 0),
            float(r["deposits_usd"] or 0),
            float(r["withdrawals_usd"] or 0),
            float(r["net_usd"] or 0),
            round(actual, 2), expected, round(actual - expected, 2),
            int(r["tx_count"] or 0),
        ])

    # ── Sheet 6: Bank Settlements ───────────────────────────────────────────
    ws6 = wb.create_sheet("Bank Settlements")
    ws6.append([
        "Date", "Amount", "Currency", "Reference", "Description",
        "Type", "Bank", "Account", "Match Status",
        "Matched CRM ID", "Confidence"
    ])
    try:
        bank_txns = queries.bank_transactions_for_period(year, mon)
        for b in bank_txns:
            ws6.append([
                str(b.get("tx_date") or ""),
                float(b.get("amount") or 0),
                b.get("acct_currency") or b.get("currency") or "",
                b.get("reference") or "",
                b.get("description") or "",
                b.get("tx_type") or "",
                b.get("bank_name") or "",
                b.get("account_number") or "",
                b.get("match_status") or "unmatched",
                b.get("matched_crm_id") or "",
                round(b.get("match_confidence") or 0, 2) if b.get("match_confidence") else "",
            ])
    except Exception:
        pass  # bank data is optional

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"recon_{month}_MRS.xlsx",
    )


@app.route("/recon/<month>/<int:login>")
@require_recon_auth
def login_detail(month, login):
    try:
        year, mon = int(month[:4]), int(month[5:7])
    except (ValueError, IndexError):
        abort(400)
    crm_rows  = queries.login_detail(year, mon, login)
    mt4_rows  = queries.login_mt4_detail(year, mon, login)
    months    = queries.available_months()
    return render_template("detail.html", month=month, login=login,
                           crm_rows=crm_rows, mt4_rows=mt4_rows, months=months)


# ---------------------------------------------------------------------------
# PSP Fee Management
# ---------------------------------------------------------------------------

FEE_TYPES = [
    "Deposit", "Withdrawal", "Settlement", "Chargeback", "Refund",
    "Rolling Reserve", "Holdback", "Setup", "Registration", "Minimum Monthly",
]

PAYMENT_METHODS = [
    "Credit Cards", "Bank Wire", "Mobile Money", "Electronic Payment",
    "Crypto", "MOMO", "E-Wallet",
]

CURRENCIES = [
    "USD", "EUR", "GBP", "ZAR", "NGN", "KES", "GHS", "UGX", "TZS",
    "RWF", "XOF", "XAF", "AED", "BRL", "CLP", "COP", "MXN", "PEN",
]

@app.route("/fees/processor-map")
@require_fees_auth
def fees_processor_map():
    """Page showing all detected Praxis processors and their agreement mappings."""
    import datetime as _dt
    # Get all distinct processors seen in last 1Y
    try:
        from db import praxis as praxis_ctx
        def _fetch():
            with praxis_ctx() as cur:
                cur.execute("""
                    SELECT payment_processor,
                           COUNT(*) AS tx_count,
                           SUM(usd_amount) AS volume,
                           SUM(fee/100.0) AS actual_fees
                    FROM praxis_transactions
                    WHERE created_timestamp >= EXTRACT(EPOCH FROM (NOW()-INTERVAL '1 year'))
                    GROUP BY payment_processor
                    ORDER BY volume DESC NULLS LAST
                """)
                return [dict(r) for r in cur.fetchall()]
        processors = queries._db_retry(_fetch)
    except Exception:
        processors = []

    agreements = queries.get_all_agreements()
    agr_safe = [{"id": a["id"], "psp_name": a["psp_name"],
                 "provider_name": a.get("provider_name") or ""}
                for a in agreements]
    proc_mappings   = queries.get_processor_mappings()
    method_mappings = queries.get_method_mappings()

    # Distinct Praxis methods in last 1Y
    try:
        from db import praxis as praxis_ctx
        def _fetch_methods():
            with praxis_ctx() as cur:
                cur.execute("""
                    SELECT payment_method,
                           COUNT(*) AS tx_count,
                           SUM(usd_amount) AS volume
                    FROM praxis_transactions
                    WHERE created_timestamp >= EXTRACT(EPOCH FROM (NOW()-INTERVAL '1 year'))
                      AND payment_method IS NOT NULL AND payment_method != ''
                    GROUP BY payment_method ORDER BY volume DESC NULLS LAST
                """)
                return [dict(r) for r in cur.fetchall()]
        methods = queries._db_retry(_fetch_methods)
    except Exception:
        methods = []

    from ai_parse import PAYMENT_METHODS as _canonical
    return render_template("fee_processor_map.html",
                           processors=processors, agreements=agr_safe,
                           proc_mappings=proc_mappings,
                           methods=methods, method_mappings=method_mappings,
                           canonical_methods=_canonical)


@app.route("/fees/method-map/save", methods=["POST"])
@require_fees_auth
def fees_method_map_save():
    data = request.get_json(force=True)
    praxis_method = (data.get("praxis_method") or "").strip()
    canonical     = (data.get("canonical") or "").strip()
    if not praxis_method:
        return jsonify({"error": "praxis_method required"}), 400
    if canonical:
        queries.save_method_mapping(praxis_method, canonical, confirmed=True)
    else:
        queries.delete_method_mapping(praxis_method)
    for k in [k for k in queries._CACHE if k.startswith("fee_calc:")]:
        del queries._CACHE[k]
    return jsonify({"ok": True})


@app.route("/fees/method-map/auto-match", methods=["POST"])
@require_fees_auth
def fees_method_map_auto():
    """AI suggests canonical names for unmapped Praxis payment methods."""
    import json as _json
    data    = request.get_json(force=True)
    methods = data.get("methods", [])
    canonical_list = ["Credit Cards","Bank Wire","Mobile Money",
                      "Electronic Payment","Crypto","MOMO","E-Wallet"]
    prompt = (
        "Map Praxis payment method names to one of these canonical names used in fee agreements:\n"
        + _json.dumps(canonical_list) + "\n\n"
        "PRAXIS METHODS TO MAP:\n" + _json.dumps(methods) + "\n\n"
        "Return ONLY a JSON array: [{\"praxis_method\": \"...\", \"canonical\": \"...\", "
        "\"confidence\": \"high|medium|low\", \"reason\": \"...\"}]\n"
        "Use null canonical if no reasonable match. Base your judgment on what the name implies "
        "(e.g. altbankonline→Electronic Payment, mobileafrica→Mobile Money, altcrypto→Crypto)."
    )
    try:
        import ai_parse as _ai
        raw = _ai._call_claude("You are a payment method name mapping assistant.", prompt)
        suggestions = _ai._safe_json_loads(raw)
        if not isinstance(suggestions, list):
            suggestions = suggestions.get("suggestions", suggestions) if isinstance(suggestions, dict) else []
        return jsonify({"suggestions": suggestions})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/fees/processor-map/save", methods=["POST"])
@require_fees_auth
def fees_processor_map_save():
    data = request.get_json(force=True)
    processor    = (data.get("processor") or "").strip()
    agreement_id = data.get("agreement_id")   # None = unmap
    if not processor:
        return jsonify({"error": "processor required"}), 400
    if agreement_id:
        queries.save_processor_mapping(processor, int(agreement_id), confirmed=True)
    else:
        queries.delete_processor_mapping(processor)
    # Bust fee_calculator cache
    for k in [k for k in queries._CACHE if k.startswith("fee_calc:")]:
        del queries._CACHE[k]
    return jsonify({"ok": True})


@app.route("/fees/processor-map/auto-match", methods=["POST"])
@require_fees_auth
def fees_processor_map_auto():
    """AI suggests agreement matches for unmapped processors."""
    import json as _json
    data        = request.get_json(force=True)
    processors  = data.get("processors", [])   # list of {name, volume}
    agreements  = queries.get_all_agreements()
    agr_list    = [{"id": a["id"], "psp_name": a["psp_name"],
                    "provider_name": a.get("provider_name") or ""} for a in agreements]

    prompt = (
        "You are matching Praxis payment processor names to PSP fee agreement names.\n\n"
        "AGREEMENTS:\n" + _json.dumps(agr_list, indent=2) + "\n\n"
        "PRAXIS PROCESSORS TO MATCH:\n" +
        _json.dumps(processors, indent=2) + "\n\n"
        "Return ONLY a JSON array of objects: "
        '[{"processor": "...", "agreement_id": <int or null>, "confidence": "high|medium|low", "reason": "..."}]\n'
        "Match by name similarity. Use null agreement_id if no reasonable match exists."
    )
    try:
        import ai_parse as _ai
        raw = _ai._call_claude("You are a payment processor name matching assistant.", prompt)
        suggestions = _ai._safe_json_loads(raw)
        if not isinstance(suggestions, list):
            suggestions = suggestions.get("suggestions", suggestions) if isinstance(suggestions, dict) else []
        return jsonify({"suggestions": suggestions})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/fees/calculator")
@require_fees_auth
def fees_calculator():
    import datetime as _dt
    span = request.args.get("span", "1m")
    span_map = {"1w":7,"1m":31,"3m":92,"6m":183,"1y":365}
    span_labels = {"1w":"1 Week","1m":"1 Month","3m":"3 Months","6m":"6 Months","1y":"1 Year"}
    today = _dt.date.today()
    date_from = today - _dt.timedelta(days=span_map.get(span,31))
    date_to   = today + _dt.timedelta(days=1)
    try:
        data = queries.fee_calculator(date_from, date_to)
    except Exception as e:
        data = {"by_psp":[],"by_method":[],"totals":{},"unmatched_processors":[],"date_from":str(date_from),"date_to":str(date_to)}
    return render_template("fee_calculator.html", data=data, span=span,
                           span_label=span_labels.get(span,span))


@app.route("/fees/calculator/uncovered")
@require_fees_auth
def fees_calculator_uncovered():
    import datetime as _dt
    processor = request.args.get("processor", "").strip()
    span      = request.args.get("span", "1m")
    span_map  = {"1w":7,"1m":31,"3m":92,"6m":183,"1y":365}
    today     = _dt.date.today()
    date_from = today - _dt.timedelta(days=span_map.get(span, 31))
    date_to   = today + _dt.timedelta(days=1)
    if not processor:
        return jsonify({"error": "processor param required"}), 400
    try:
        rows = queries.fee_uncovered_transactions(processor, date_from, date_to)
        return jsonify({"processor": processor, "span": span, "transactions": rows})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/fees/db-backup")
@require_fees_auth
def fees_db_backup():
    """Legacy URL — redirect to admin panel."""
    return redirect(url_for("fees_admin"))


# ---------------------------------------------------------------------------
# Admin panel
# ---------------------------------------------------------------------------

_SNAPSHOTS_DIR = os.path.join(os.path.dirname(__file__), "db_snapshots")
_FEES_DB_PATH  = os.path.join(os.path.dirname(__file__), "fees.db")


def _snapshot_label_safe(label):
    import re as _re
    return bool(_re.match(r'^[A-Za-z0-9_\-]{1,40}$', label or ""))


def _list_snapshots():
    import os as _os
    if not _os.path.isdir(_SNAPSHOTS_DIR):
        return []
    snaps = []
    for fn in sorted(_os.listdir(_SNAPSHOTS_DIR)):
        if fn.endswith(".db"):
            path = _os.path.join(_SNAPSHOTS_DIR, fn)
            st   = _os.stat(path)
            snaps.append({
                "name":     fn[:-3],
                "filename": fn,
                "size_mb":  round(st.st_size / 1_048_576, 1),
                "modified": __import__("datetime").datetime.fromtimestamp(st.st_mtime).strftime("%Y-%m-%d %H:%M"),
            })
    return snaps


@app.route("/fees/admin")
@require_admin_auth
def fees_admin():
    import os as _os, datetime as _dt
    from db import FEES_MODE
    db_stat = _os.stat(_FEES_DB_PATH) if _os.path.exists(_FEES_DB_PATH) else None
    db_info = {
        "size_mb":  round(db_stat.st_size / 1_048_576, 1) if db_stat else 0,
        "modified": _dt.datetime.fromtimestamp(db_stat.st_mtime).strftime("%Y-%m-%d %H:%M") if db_stat else "—",
    } if db_stat else None
    return render_template("fees_admin.html",
                           fees_mode=FEES_MODE,
                           db_info=db_info,
                           snapshots=_list_snapshots())


@app.route("/fees/admin/backup")
@require_admin_auth
def fees_admin_backup():
    from datetime import date as _d
    if not os.path.exists(_FEES_DB_PATH):
        abort(404, "fees.db not found")
    return send_file(_FEES_DB_PATH, as_attachment=True,
                     download_name=f"fees_backup_{_d.today()}.db",
                     mimetype="application/octet-stream")


@app.route("/fees/admin/upload", methods=["POST"])
@require_admin_auth
def fees_admin_upload():
    from db import FEES_MODE
    if FEES_MODE != "demo":
        return jsonify({"ok": False, "error": "Only available in DEMO mode"}), 400
    f = request.files.get("db_file")
    if not f or not f.filename.endswith(".db"):
        return jsonify({"ok": False, "error": "Please upload a .db file"}), 400
    data = f.read()
    # Validate it's a SQLite file (magic bytes)
    if not data.startswith(b"SQLite format 3\x00"):
        return jsonify({"ok": False, "error": "File is not a valid SQLite database"}), 400
    # Bust all caches before replacing
    queries._CACHE.clear()
    with open(_FEES_DB_PATH, "wb") as fh:
        fh.write(data)
    return jsonify({"ok": True})


@app.route("/fees/admin/snapshot/save", methods=["POST"])
@require_admin_auth
def fees_admin_snapshot_save():
    import shutil as _sh
    label = request.form.get("label", "").strip()
    if not _snapshot_label_safe(label):
        return jsonify({"ok": False, "error": "Label must be 1–40 alphanumeric/dash/underscore chars"}), 400
    if not os.path.exists(_FEES_DB_PATH):
        return jsonify({"ok": False, "error": "fees.db not found"}), 404
    os.makedirs(_SNAPSHOTS_DIR, exist_ok=True)
    dest = os.path.join(_SNAPSHOTS_DIR, f"{label}.db")
    _sh.copy2(_FEES_DB_PATH, dest)
    return jsonify({"ok": True})


@app.route("/fees/admin/snapshot/<name>/activate", methods=["POST"])
@require_admin_auth
def fees_admin_snapshot_activate(name):
    import shutil as _sh
    from db import FEES_MODE
    if FEES_MODE != "demo":
        return jsonify({"ok": False, "error": "Only available in DEMO mode"}), 400
    if not _snapshot_label_safe(name):
        abort(400)
    src = os.path.join(_SNAPSHOTS_DIR, f"{name}.db")
    if not os.path.exists(src):
        abort(404)
    queries._CACHE.clear()
    _sh.copy2(src, _FEES_DB_PATH)
    return jsonify({"ok": True})


@app.route("/fees/admin/snapshot/<name>/download")
@require_admin_auth
def fees_admin_snapshot_download(name):
    if not _snapshot_label_safe(name):
        abort(400)
    path = os.path.join(_SNAPSHOTS_DIR, f"{name}.db")
    if not os.path.exists(path):
        abort(404)
    return send_file(path, as_attachment=True, download_name=f"{name}.db",
                     mimetype="application/octet-stream")


@app.route("/fees/admin/snapshot/<name>/delete", methods=["POST"])
@require_admin_auth
def fees_admin_snapshot_delete(name):
    if not _snapshot_label_safe(name):
        abort(400)
    path = os.path.join(_SNAPSHOTS_DIR, f"{name}.db")
    if os.path.exists(path):
        os.remove(path)
    return jsonify({"ok": True})


@app.route("/fees/mode", methods=["POST"])
@require_admin_auth
def fees_switch_mode():
    """Switch FEES_MODE between demo and live by updating docker-compose.yml
    and restarting the container. Server-only — no-op in local dev."""
    import os as _os, subprocess as _sp
    new_mode = request.form.get("mode", "demo").lower()
    if new_mode not in ("demo", "live"):
        abort(400, "mode must be demo or live")

    compose_path = "/root/recon-app/docker-compose.yml"
    if not _os.path.exists(compose_path):
        # Local dev — just show a message
        return jsonify({"ok": False, "msg": "compose file not found (local dev?)"}), 400

    try:
        with open(compose_path) as f:
            content = f.read()
        import re as _re
        content = _re.sub(r'FEES_MODE=\w+', f'FEES_MODE={new_mode}', content)
        with open(compose_path, "w") as f:
            f.write(content)
        # Restart the container to pick up the new env var
        _sp.run(["docker", "restart", "recon-app-recon-1"],
                capture_output=True, timeout=30)
        return jsonify({"ok": True, "mode": new_mode})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500


@app.route("/fees")
@require_fees_auth
def fees_list():
    agreements = queries.get_all_agreements()
    terminated = queries.get_terminated_agreements()
    entities = queries.get_entities()
    from db import FEES_MODE
    return render_template("fees.html", agreements=agreements, terminated=terminated,
                           entities=entities, fees_mode=FEES_MODE)


@app.route("/fees/entities/add", methods=["POST"])
@require_fees_auth
def fees_entity_add():
    name = request.form.get("name", "").strip().upper()
    if name:
        queries.add_entity(name)
    return redirect(url_for("fees_list"))


@app.route("/fees/entities/delete", methods=["POST"])
@require_fees_auth
def fees_entity_delete():
    name = request.form.get("name", "")
    queries.delete_entity(name)
    return redirect(url_for("fees_list"))


@app.route("/fees/new", methods=["GET", "POST"])
@require_fees_auth
def fees_new():
    if request.method == "POST":
        data = {
            "psp_name":         request.form["psp_name"].strip(),
            "provider_name":    request.form.get("provider_name", "").strip() or None,
            "agreement_entity": request.form.get("agreement_entity", "").strip() or None,
            "agreement_date":   request.form.get("agreement_date") or None,
            "addendum_date":    request.form.get("addendum_date") or None,
            "auto_settlement":  request.form.get("auto_settlement") == "on",
            "settlement_bank":  request.form.get("settlement_bank", "").strip() or None,
        }
        if not data["psp_name"]:
            abort(400, "PSP name is required")
        psp_id = queries.create_agreement(data)
        return redirect(url_for("fees_detail", psp_id=psp_id))
    return render_template("fee_form.html", agreement=None, entities=queries.get_entities(),
                           today=date.today().isoformat())


@app.route("/fees/<int:psp_id>")
@require_fees_auth
def fees_detail(psp_id):
    agreement = queries.get_agreement(psp_id)
    if not agreement:
        abort(404)
    rules = queries.get_fee_rules(psp_id)
    return render_template("fee_detail.html", agreement=agreement, rules=rules,
                           fee_types=FEE_TYPES, payment_methods=PAYMENT_METHODS,
                           currencies=CURRENCIES, entities=queries.get_entities())


@app.route("/fees/<int:psp_id>/edit", methods=["POST"])
@require_fees_auth
def fees_edit(psp_id):
    data = {
        "psp_name":         request.form["psp_name"].strip(),
        "provider_name":    request.form.get("provider_name", "").strip() or None,
        "agreement_entity": request.form.get("agreement_entity", "").strip() or None,
        "agreement_date":   request.form.get("agreement_date") or None,
        "addendum_date":    request.form.get("addendum_date") or None,
        "auto_settlement":  request.form.get("auto_settlement") == "on",
        "settlement_bank":  request.form.get("settlement_bank", "").strip() or None,
    }
    queries.update_agreement(psp_id, data)
    return redirect(url_for("fees_detail", psp_id=psp_id))


@app.route("/fees/add")
@require_fees_auth
def fees_add_manual():
    """Render the confirm form with empty data for manual agreement entry."""
    historical = request.args.get("historical") == "1"
    return render_template(
        "fee_confirm.html",
        agreement={}, fee_rules=[], dups_removed=0,
        ai_warnings=[], raw_response="", filename=None,
        cache_token=None, historical=historical, manual=True,
        entities=queries.get_entities(),
        fee_types=FEE_TYPES, payment_methods=PAYMENT_METHODS, currencies=CURRENCIES,
        gaps=[],
    )


@app.route("/fees/<int:psp_id>/delete", methods=["POST"])
@require_fees_auth
def fees_delete(psp_id):
    queries.delete_agreement(psp_id)
    return redirect(url_for("fees_list"))


@app.route("/fees/<int:psp_id>/purge", methods=["POST"])
@require_fees_auth
def fees_purge(psp_id):
    """Permanently delete a terminated agreement from history."""
    queries.purge_agreement(psp_id)
    return jsonify({"ok": True})


@app.route("/fees/<int:psp_id>/rules/add", methods=["POST"])
@require_fees_auth
def fees_add_rule(psp_id):
    kind = request.form["fee_kind"]
    data = {
        "payment_method": request.form.get("payment_method", "").strip() or None,
        "fee_type":       request.form["fee_type"],
        "country":        request.form.get("country", "GLOBAL").strip() or "GLOBAL",
        "sub_provider":   request.form.get("sub_provider", "").strip() or None,
        "fee_kind":       kind,
        "pct_rate":       None,
        "fixed_amount":   None,
        "fixed_currency": None,
        "description":    request.form.get("description", "").strip() or None,
        "tiers":          [],
    }

    if kind in ("percentage", "fixed_plus_pct"):
        raw = request.form.get("pct_rate", "")
        if raw:
            data["pct_rate"] = float(raw) / 100.0

    if kind in ("fixed", "fixed_plus_pct"):
        raw = request.form.get("fixed_amount", "")
        if raw:
            data["fixed_amount"] = float(raw)
        data["fixed_currency"] = request.form.get("fixed_currency", "USD")

    if kind == "tiered":
        froms = request.form.getlist("tier_from")
        tos   = request.form.getlist("tier_to")
        rates = request.form.getlist("tier_rate")
        for f, t, r in zip(froms, tos, rates):
            if r:
                data["tiers"].append({
                    "volume_from": float(f) if f else 0,
                    "volume_to":   float(t) if t else None,
                    "pct_rate":    float(r) / 100.0,
                })

    queries.create_fee_rule(psp_id, data)
    return redirect(url_for("fees_detail", psp_id=psp_id))


@app.route("/fees/rules/<int:rule_id>/edit", methods=["POST"])
@require_fees_auth
def fees_edit_rule(rule_id):
    psp_id = request.form.get("psp_id", type=int)
    kind = request.form["fee_kind"]
    data = {
        "payment_method": request.form.get("payment_method", "").strip() or None,
        "fee_type":       request.form["fee_type"],
        "country":        request.form.get("country", "GLOBAL").strip() or "GLOBAL",
        "sub_provider":   request.form.get("sub_provider", "").strip() or None,
        "fee_kind":       kind,
        "pct_rate":       None,
        "fixed_amount":   None,
        "fixed_currency": None,
        "description":    request.form.get("description", "").strip() or None,
        "tiers":          [],
    }
    if kind in ("percentage", "fixed_plus_pct"):
        raw = request.form.get("pct_rate", "")
        if raw:
            data["pct_rate"] = float(raw) / 100.0
    if kind in ("fixed", "fixed_plus_pct"):
        raw = request.form.get("fixed_amount", "")
        if raw:
            data["fixed_amount"] = float(raw)
        data["fixed_currency"] = request.form.get("fixed_currency", "USD")
    if kind == "tiered":
        froms = request.form.getlist("tier_from")
        tos   = request.form.getlist("tier_to")
        rates = request.form.getlist("tier_rate")
        for f, t, r in zip(froms, tos, rates):
            if r:
                data["tiers"].append({
                    "volume_from": float(f) if f else 0,
                    "volume_to":   float(t) if t else None,
                    "pct_rate":    float(r) / 100.0,
                })
    queries.update_fee_rule(rule_id, data)
    return redirect(url_for("fees_detail", psp_id=psp_id))


@app.route("/fees/rules/<int:rule_id>/delete", methods=["POST"])
@require_fees_auth
def fees_delete_rule(rule_id):
    psp_id = request.form.get("psp_id", type=int)
    queries.delete_fee_rule(rule_id)
    return redirect(url_for("fees_detail", psp_id=psp_id))


# ---------------------------------------------------------------------------
# AI Agreement Upload
# ---------------------------------------------------------------------------

ALLOWED_UPLOAD_EXTS = {"pdf", "docx", "doc"}


@app.route("/fees/upload", methods=["GET", "POST"])
@require_fees_auth
def fees_upload():
    templates     = queries.get_prompt_templates()
    context_notes = queries.get_context_notes()

    # Amendment mode: psp_id present → amending an existing agreement
    psp_id_get  = request.args.get("psp_id", type=int)
    amend_agr   = queries.get_agreement(psp_id_get) if psp_id_get else None

    # Choose the default template for the current mode (each type has its own default)
    if amend_agr:
        default_tpl = next((t for t in templates if t.get("prompt_type") == "amendment" and t["is_default"]), None)
        if not default_tpl:
            default_tpl = next((t for t in templates if t.get("prompt_type") == "amendment"), None)
    else:
        default_tpl = next((t for t in templates if t.get("prompt_type", "agreement") == "agreement" and t["is_default"]), None)
        if not default_tpl:
            default_tpl = next((t for t in templates if t["is_default"]), None)
    if not default_tpl and templates:
        default_tpl = templates[0]

    historical = (request.args.get("historical") == "1"
                  if request.method == "GET"
                  else request.form.get("historical") == "1")

    if request.method == "GET":
        return render_template("fee_upload.html", templates=templates,
                               default_tpl=default_tpl, context_notes=context_notes,
                               amendment_agreement=amend_agr, historical=historical)

    def _render_error(msg):
        psp_id_post = request.form.get("psp_id", type=int)
        agr = queries.get_agreement(psp_id_post) if psp_id_post else None
        return render_template("fee_upload.html", templates=templates,
                               default_tpl=default_tpl, context_notes=context_notes,
                               amendment_agreement=agr, historical=historical, error=msg)

    f = request.files.get("agreement_file")
    if not f or not f.filename:
        return _render_error("No file selected.")

    ext = f.filename.rsplit(".", 1)[-1].lower() if "." in f.filename else ""
    if ext not in ALLOWED_UPLOAD_EXTS:
        return _render_error("Only PDF and DOCX files are supported.")

    # Resolve selected prompt template
    tpl_id = request.form.get("prompt_template_id", type=int)
    tpl    = queries.get_prompt_template(tpl_id) if tpl_id else default_tpl
    system_prompt = tpl["system_prompt"] if tpl else None

    # Collect all extra context: checked saved notes + newly typed boxes
    extra_sections = []

    checked_ids = request.form.getlist("saved_note_ids", type=int)
    if checked_ids:
        queries.increment_context_note_usage(checked_ids)
        saved = {n["id"]: n for n in context_notes}
        for nid in checked_ids:
            if nid in saved:
                extra_sections.append((saved[nid]["label"], saved[nid]["text"]))

    ctx_labels     = request.form.getlist("ctx_label")
    ctx_texts      = request.form.getlist("ctx_text")
    ctx_saves      = request.form.getlist("ctx_save")       # "1" or "" per box
    ctx_note_types = request.form.getlist("ctx_note_type")  # "agreement" or "amendment" per box
    for i, (label, text) in enumerate(zip(ctx_labels, ctx_texts)):
        text = text.strip()
        if not text:
            continue
        label     = label.strip() or "Additional Note"
        note_type = ctx_note_types[i] if i < len(ctx_note_types) else "agreement"
        extra_sections.append((label, text))
        if i < len(ctx_saves) and ctx_saves[i] == "1":
            queries.save_context_note(label, text, note_type=note_type)

    if extra_sections:
        divider = "\n" + "═" * 55
        addon = divider + "\n  ADDITIONAL CONTEXT (provided before analysis)\n" + divider
        for label, text in extra_sections:
            addon += f"\n\n── {label} ──\n{text}"
        system_prompt = (system_prompt or "") + "\n\n" + addon

    # Amendment mode: psp_id in form → amend an existing agreement
    psp_id_post = request.form.get("psp_id", type=int)

    vision_mode = request.form.get("vision_mode") == "1"

    try:
        file_bytes = f.read()
        text  = ai_parse.extract_text(file_bytes, f.filename)
        pages = ai_parse.extract_pages(file_bytes, f.filename)
    except Exception as e:
        return _render_error(f"Could not read file: {e}")

    if not text.strip():
        if not vision_mode or ext != "pdf":
            return _render_error(
                "No text could be extracted from this file. "
                "If this is a scanned PDF, tick \"Scanned PDF (vision)\" and re-upload.")

    # ── New agreement flow: save file alongside the agreement ───────────────
    if not psp_id_post:
        agr_cache_token = str(uuid.uuid4())
        queries.cache_upload(agr_cache_token, f.filename, file_bytes)
    else:
        agr_cache_token = None

    # ── Amendment flow ──────────────────────────────────────────────────────
    if psp_id_post:
        amendment_agreement = queries.get_agreement(psp_id_post)
        if not amendment_agreement:
            return _render_error("PSP not found.")

        try:
            if vision_mode:
                result = ai_parse.analyze_amendment_vision(file_bytes, f.filename,
                                                           system_prompt=system_prompt)
            else:
                result = ai_parse.analyze_amendment(text, system_prompt=system_prompt, pages=pages)
        except Exception as e:
            return _render_error(f"AI analysis failed: {e}")

        raw_changes   = result.get("rule_changes", [])
        amend_meta    = result.get("amendment", {})
        addendum_date = amend_meta.get("addendum_date") or ""
        notes         = amend_meta.get("notes") or ""

        # Cache file so confirm route can persist it
        cache_token = str(uuid.uuid4())
        queries.cache_upload(cache_token, f.filename, file_bytes)

        existing_rules = queries.get_fee_rules(psp_id_post)

        def _find_match(match_on):
            pm = (match_on.get("payment_method") or "").lower()
            co = (match_on.get("country") or "").lower()
            ft = (match_on.get("fee_type") or "").lower()
            for r in existing_rules:
                if (
                    (r.get("payment_method") or "").lower() == pm
                    and (r.get("country") or "").lower() == co
                    and (r.get("fee_type") or "").lower() == ft
                ):
                    return r
            return None

        changes = []
        for rc in raw_changes:
            action   = rc.get("action", "add")
            match_on = rc.get("match_on") or {}
            existing = None

            if action in ("replace", "remove"):
                existing = _find_match(match_on)
                if existing is None:
                    action = "add"   # can't find what to replace — treat as new

            # Cross-session duplicate check: for "add" actions, also verify the rule
            # doesn't already exist in the DB (e.g. same amendment uploaded twice,
            # or two amendments covering the same jurisdiction).
            # If a match is found, downgrade to "replace" so the user sees the conflict
            # and can decide to skip or overwrite — instead of silently creating a duplicate.
            auto_replaced = False
            if action == "add":
                existing = _find_match({
                    "payment_method": rc.get("payment_method"),
                    "country":        rc.get("country"),
                    "fee_type":       rc.get("fee_type"),
                })
                if existing is not None:
                    action        = "replace"
                    auto_replaced = True   # flag so UI can warn the user

            changes.append({
                "action":        action,
                "new_rule":      rc,
                "existing_rule": existing,
                "auto_replaced": auto_replaced,
            })

        return render_template(
            "fee_amendment_confirm.html",
            agreement=amendment_agreement,
            changes=changes,
            addendum_date=addendum_date,
            notes=notes,
            cache_token=cache_token,
            dups_removed=result.get("dups_removed", 0),
            ai_warnings=result.get("warnings", []),
            raw_response=result.get("raw_response", ""),
            fee_types=FEE_TYPES,
            payment_methods=PAYMENT_METHODS,
            currencies=CURRENCIES,
            filename=f.filename,
        )

    # ── New agreement flow ─────────────────────────────────────────────────
    try:
        if vision_mode:
            result = ai_parse.analyze_agreement_vision(file_bytes, f.filename,
                                                       system_prompt=system_prompt)
        else:
            result = ai_parse.analyze_agreement(text, system_prompt=system_prompt, pages=pages)
    except Exception as e:
        return _render_error(f"AI analysis failed: {e}")

    return render_template(
        "fee_confirm.html",
        agreement=result.get("agreement", {}),
        fee_rules=result.get("fee_rules", []),
        dups_removed=result.get("dups_removed", 0),
        ai_warnings=result.get("warnings", []),
        raw_response=result.get("raw_response", ""),
        entities=queries.get_entities(),
        fee_types=FEE_TYPES,
        payment_methods=PAYMENT_METHODS,
        currencies=CURRENCIES,
        filename=f.filename,
        cache_token=agr_cache_token,
        historical=historical,
    )


@app.route("/fees/upload/confirm", methods=["POST"])
@require_fees_auth
def fees_upload_confirm():
    data = {
        "psp_name":         request.form["psp_name"].strip(),
        "provider_name":    request.form.get("provider_name", "").strip() or None,
        "agreement_entity": request.form.get("agreement_entity", "").strip() or None,
        "agreement_date":   request.form.get("agreement_date") or None,
        "addendum_date":    request.form.get("addendum_date") or None,
        "auto_settlement":  request.form.get("auto_settlement") == "on",
        "settlement_bank":  request.form.get("settlement_bank", "").strip() or None,
    }
    if not data["psp_name"]:
        abort(400, "PSP name is required")

    psp_id = queries.create_agreement(data)

    rule_count = int(request.form.get("rule_count", 0))
    for i in range(rule_count):
        kind = request.form.get(f"fee_kind_{i}", "percentage")
        rule = {
            "payment_method": request.form.get(f"payment_method_{i}", "").strip() or None,
            "fee_type":       request.form.get(f"fee_type_{i}", "Deposit"),
            "country":        request.form.get(f"country_{i}", "GLOBAL").strip() or "GLOBAL",
            "sub_provider":   request.form.get(f"sub_provider_{i}", "").strip() or None,
            "fee_kind":       kind,
            "pct_rate":       None,
            "fixed_amount":   None,
            "fixed_currency": None,
            "description":    request.form.get(f"description_{i}", "").strip() or None,
            "tiers":          [],
        }
        if kind in ("percentage", "fixed_plus_pct"):
            raw = request.form.get(f"pct_rate_{i}", "")
            if raw:
                rule["pct_rate"] = float(raw) / 100.0
        if kind in ("fixed", "fixed_plus_pct"):
            raw = request.form.get(f"fixed_amount_{i}", "")
            if raw:
                rule["fixed_amount"] = float(raw)
            rule["fixed_currency"] = request.form.get(f"fixed_currency_{i}", "USD")
        if kind == "tiered":
            froms = request.form.getlist(f"tier_from_{i}")
            tos   = request.form.getlist(f"tier_to_{i}")
            rates = request.form.getlist(f"tier_rate_{i}")
            for f_v, t_v, r_v in zip(froms, tos, rates):
                if r_v:
                    rule["tiers"].append({
                        "volume_from": float(f_v) if f_v else 0,
                        "volume_to":   float(t_v) if t_v else None,
                        "pct_rate":    float(r_v) / 100.0,
                    })
        queries.create_fee_rule(psp_id, rule)

    # Save initial agreement file if a cache token was passed
    cache_token = request.form.get("cache_token", "").strip()
    if cache_token:
        fn, fd = queries.pop_upload_cache(cache_token)
        if fn and fd:
            queries.save_agreement_file(psp_id, fn, fd)

    # Historical mode: immediately deactivate so it lands in the Historical tab
    if request.form.get("historical") == "1":
        queries.delete_agreement(psp_id)

    return redirect(url_for("fees_detail", psp_id=psp_id))


@app.route("/fees/<int:psp_id>/amendment/confirm", methods=["POST"])
@require_fees_auth
def fees_amendment_confirm(psp_id):
    """Apply AI-extracted amendment changes and record them in amendment history."""
    if not queries.get_agreement(psp_id):
        abort(404)

    cache_token   = request.form.get("cache_token", "").strip()
    addendum_date = (request.form.get("addendum_date") or "").strip()
    notes         = (request.form.get("notes") or "").strip()
    change_count  = int(request.form.get("change_count", 0))

    # Retrieve file from cache
    filename, file_data = queries.pop_upload_cache(cache_token) if cache_token else (None, None)

    applied_changes = []  # list of (action, fee_rule_id, old_rule, new_rule)

    for i in range(change_count):
        if request.form.get(f"skip_{i}"):
            continue

        action = request.form.get(f"action_{i}", "add")

        if action == "remove":
            rule_id = request.form.get(f"existing_rule_id_{i}", type=int)
            old_rule = queries.get_fee_rule(rule_id) if rule_id else None
            if rule_id:
                queries.delete_fee_rule(rule_id)
            applied_changes.append((action, rule_id, old_rule, None))
            continue

        kind = request.form.get(f"fee_kind_{i}", "percentage")
        rule = {
            "payment_method": request.form.get(f"payment_method_{i}", "").strip() or None,
            "fee_type":       request.form.get(f"fee_type_{i}", "Deposit"),
            "country":        request.form.get(f"country_{i}", "GLOBAL").strip() or "GLOBAL",
            "sub_provider":   request.form.get(f"sub_provider_{i}", "").strip() or None,
            "fee_kind":       kind,
            "pct_rate":       None,
            "fixed_amount":   None,
            "fixed_currency": None,
            "description":    request.form.get(f"description_{i}", "").strip() or None,
            "tiers":          [],
        }
        if kind in ("percentage", "fixed_plus_pct"):
            raw = request.form.get(f"pct_rate_{i}", "")
            if raw:
                rule["pct_rate"] = float(raw) / 100.0
        if kind in ("fixed", "fixed_plus_pct"):
            raw = request.form.get(f"fixed_amount_{i}", "")
            if raw:
                rule["fixed_amount"] = float(raw)
            rule["fixed_currency"] = request.form.get(f"fixed_currency_{i}", "USD")
        if kind == "tiered":
            froms = request.form.getlist(f"tier_from_{i}")
            tos   = request.form.getlist(f"tier_to_{i}")
            rates = request.form.getlist(f"tier_rate_{i}")
            for f_v, t_v, r_v in zip(froms, tos, rates):
                if r_v:
                    rule["tiers"].append({
                        "volume_from": float(f_v) if f_v else 0,
                        "volume_to":   float(t_v) if t_v else None,
                        "pct_rate":    float(r_v) / 100.0,
                    })

        if action == "replace":
            rule_id = request.form.get(f"existing_rule_id_{i}", type=int)
            old_rule = queries.get_fee_rule(rule_id) if rule_id else None
            if rule_id:
                queries.update_fee_rule(rule_id, rule)
                result_id = rule_id
            else:
                result_id = queries.create_fee_rule(psp_id, rule)
            applied_changes.append((action, result_id, old_rule, rule))
        else:
            result_id = queries.create_fee_rule(psp_id, rule)
            applied_changes.append((action, result_id, None, rule))

    if addendum_date:
        queries.update_addendum_date(psp_id, addendum_date)

    # Record amendment in history
    amend_id = queries.create_amendment_record(
        psp_id, addendum_date or None, filename, file_data,
        notes or None, len(applied_changes)
    )
    for action, fee_rule_id, old_rule, new_rule in applied_changes:
        queries.add_amendment_change(amend_id, action, fee_rule_id, old_rule, new_rule)

    return redirect(url_for("fees_detail", psp_id=psp_id))


# ---------------------------------------------------------------------------
# Amendment History & File Downloads
# ---------------------------------------------------------------------------

@app.route("/fees/<int:psp_id>/amendments")
@require_fees_auth
def fees_amendments(psp_id):
    agreement  = queries.get_agreement(psp_id)
    if not agreement:
        abort(404)
    amendments = queries.get_amendments(psp_id)
    has_agr_file = bool(queries.get_agreement_file(psp_id)[0])
    return render_template("fee_amendment_history.html",
                           agreement=agreement, amendments=amendments,
                           has_agr_file=has_agr_file)


@app.route("/fees/<int:psp_id>/amendments/<int:amend_id>")
@require_fees_auth
def fees_amendment_detail(psp_id, amend_id):
    agreement = queries.get_agreement(psp_id)
    if not agreement:
        abort(404)
    amend = queries.get_amendment(amend_id)
    if not amend or amend["agreement_id"] != psp_id:
        abort(404)
    return render_template("fee_amendment_history.html",
                           agreement=agreement,
                           amendments=queries.get_amendments(psp_id),
                           selected=amend,
                           has_agr_file=bool(queries.get_agreement_file(psp_id)[0]))


@app.route("/fees/<int:psp_id>/download-agreement")
@require_fees_auth
def fees_download_agreement(psp_id):
    filename, file_data = queries.get_agreement_file(psp_id)
    if not file_data:
        abort(404)
    mime = "application/pdf" if (filename or "").lower().endswith(".pdf") else "application/octet-stream"
    return send_file(io.BytesIO(file_data), download_name=filename, mimetype=mime, as_attachment=True)


@app.route("/fees/amendments/<int:amend_id>/download")
@require_fees_auth
def fees_download_amendment(amend_id):
    filename, file_data = queries.get_amendment_file(amend_id)
    if not file_data:
        abort(404)
    mime = "application/pdf" if (filename or "").lower().endswith(".pdf") else "application/octet-stream"
    return send_file(io.BytesIO(file_data), download_name=filename, mimetype=mime, as_attachment=True)


# ---------------------------------------------------------------------------
# Prompt Template Management
# ---------------------------------------------------------------------------

@app.route("/fees/context-notes/save", methods=["POST"])
@require_fees_auth
def context_note_save():
    data      = request.get_json(force=True)
    label     = (data.get("label")     or "").strip()
    text      = (data.get("text")      or "").strip()
    note_type = (data.get("note_type") or "agreement").strip()
    if note_type not in ("agreement", "amendment"):
        note_type = "agreement"
    if not text:
        return jsonify({"error": "Text is required"}), 400
    label = label or "Additional Note"
    note_id = queries.save_context_note(label, text, note_type=note_type)
    return jsonify({"id": note_id, "label": label, "text": text,
                    "note_type": note_type, "use_count": 0, "usage_pct": 0})


@app.route("/fees/context-notes/<int:note_id>/delete", methods=["POST"])
@require_fees_auth
def context_note_delete(note_id):
    queries.delete_context_note(note_id)
    return redirect(url_for("fees_upload"))


@app.route("/fees/prompts")
@require_fees_auth
def prompt_list():
    templates = queries.get_prompt_templates()
    return render_template("prompt_list.html", templates=templates)


@app.route("/fees/prompts/new", methods=["GET", "POST"])
@require_fees_auth
def prompt_new():
    templates = queries.get_prompt_templates()
    if request.method == "POST":
        name   = request.form.get("name", "").strip()
        prompt = request.form.get("system_prompt", "").strip()
        if not name or not prompt:
            return render_template("prompt_form.html", template=None, templates=templates,
                                   error="Name and prompt text are required.")
        prompt_type = request.form.get("prompt_type", "agreement")
        tpl_id = queries.create_prompt_template(name, prompt, prompt_type=prompt_type)
        if request.form.get("set_default"):
            queries.set_default_prompt_template(tpl_id)
        return redirect(url_for("prompt_list"))
    base_id = request.args.get("base", type=int)
    return render_template("prompt_form.html", template=None, templates=templates,
                           base_id=base_id)


@app.route("/fees/prompts/<int:tpl_id>/edit", methods=["GET", "POST"])
@require_fees_auth
def prompt_edit(tpl_id):
    tpl = queries.get_prompt_template(tpl_id)
    if not tpl:
        abort(404)
    if tpl.get("is_builtin"):
        return redirect(url_for("prompt_list"))
    if request.method == "POST":
        name   = request.form.get("name", "").strip()
        prompt = request.form.get("system_prompt", "").strip()
        if not name or not prompt:
            return render_template("prompt_form.html", template=tpl,
                                   error="Name and prompt text are required.")
        prompt_type = request.form.get("prompt_type", "agreement")
        queries.update_prompt_template(tpl_id, name, prompt, prompt_type=prompt_type)
        if request.form.get("set_default"):
            queries.set_default_prompt_template(tpl_id)
        return redirect(url_for("prompt_list"))
    return render_template("prompt_form.html", template=tpl)


@app.route("/fees/prompts/<int:tpl_id>/content")
@require_fees_auth
def prompt_content(tpl_id):
    tpl = queries.get_prompt_template(tpl_id)
    if not tpl:
        abort(404)
    return jsonify({"system_prompt": tpl["system_prompt"]})


@app.route("/fees/prompts/<int:tpl_id>/set-default", methods=["POST"])
@require_fees_auth
def prompt_set_default(tpl_id):
    queries.set_default_prompt_template(tpl_id)
    return redirect(url_for("prompt_list"))


@app.route("/fees/prompts/<int:tpl_id>/delete", methods=["POST"])
@require_fees_auth
def prompt_delete(tpl_id):
    tpl = queries.get_prompt_template(tpl_id)
    if tpl and tpl.get("is_builtin"):
        abort(403)
    queries.delete_prompt_template(tpl_id)
    return redirect(url_for("prompt_list"))


# ═══════════════════════════════════════════════════════════════════════════════
# BANK STATEMENTS MODULE
# ═══════════════════════════════════════════════════════════════════════════════

BANK_UPLOAD_EXTS = {"csv", "xls", "xlsx", "pdf"}


@app.route("/banks")
@require_recon_auth
def banks_main():
    accounts   = queries.get_bank_accounts()
    statements = queries.get_bank_statements()
    hist_accounts   = queries.get_historical_accounts()
    hist_statements = queries.get_historical_statements()
    return render_template("banks.html",
                           accounts=accounts, statements=statements,
                           hist_accounts=hist_accounts, hist_statements=hist_statements)


@app.route("/banks/accounts/add", methods=["POST"])
@require_recon_auth
def bank_account_add():
    queries.create_bank_account({
        "bank_name":      request.form["bank_name"],
        "account_number": request.form["account_number"],
        "account_label":  request.form.get("account_label", ""),
        "currency":       request.form.get("currency", "USD"),
        "entity":         request.form.get("entity", ""),
    })
    return redirect(url_for("banks_main"))


@app.route("/banks/accounts/<int:account_id>/edit", methods=["POST"])
@require_recon_auth
def bank_account_edit(account_id):
    queries.update_bank_account(account_id, {
        "bank_name":      request.form["bank_name"],
        "account_number": request.form["account_number"],
        "account_label":  request.form.get("account_label", ""),
        "currency":       request.form.get("currency", "USD"),
        "entity":         request.form.get("entity", ""),
    })
    return redirect(url_for("banks_main"))


@app.route("/banks/accounts/<int:account_id>/delete", methods=["POST"])
@require_recon_auth
def bank_account_delete(account_id):
    queries.delete_bank_account(account_id)
    return redirect(url_for("banks_main"))


@app.route("/banks/accounts/<int:account_id>/restore", methods=["POST"])
@require_recon_auth
def bank_account_restore(account_id):
    queries.restore_bank_account(account_id)
    return redirect(url_for("banks_main"))


@app.route("/banks/accounts/<int:account_id>/purge", methods=["POST"])
@require_recon_auth
def bank_account_purge(account_id):
    queries.purge_bank_account(account_id)
    return redirect(url_for("banks_main"))


@app.route("/banks/upload", methods=["GET", "POST"])
@require_recon_auth
def bank_upload():
    accounts = queries.get_bank_accounts()
    if request.method == "GET":
        return render_template("bank_upload.html", accounts=accounts)

    f = request.files.get("statement_file")
    if not f or not f.filename:
        return render_template("bank_upload.html", accounts=accounts,
                               error="No file selected.")

    ext = f.filename.rsplit(".", 1)[-1].lower() if "." in f.filename else ""
    if ext not in BANK_UPLOAD_EXTS:
        return render_template("bank_upload.html", accounts=accounts,
                               error="Supported formats: CSV, XLS, XLSX, PDF.")

    account_id = request.form.get("bank_account_id", type=int)
    account = queries.get_bank_account(account_id) if account_id else None

    vision_mode = request.form.get("vision_mode") == "1"

    try:
        file_bytes = f.read()
        import bank_parse
        result = bank_parse.parse_bank_statement(file_bytes, f.filename,
                                                  vision_mode=vision_mode)
    except Exception as e:
        return render_template("bank_upload.html", accounts=accounts,
                               error=f"Parse error: {e}")

    # Auto-match detected account number to an existing bank account
    matched_account = account
    if not matched_account:
        detected_num = (result.get("account_number") or "").strip().lstrip("0")
        if detected_num:
            for a in accounts:
                db_num = (a.get("account_number") or "").strip().lstrip("0")
                if db_num and (db_num == detected_num or db_num.endswith(detected_num) or detected_num.endswith(db_num)):
                    matched_account = a
                    break

    # Cache the raw file so confirm step can store it for later download
    file_token = str(uuid.uuid4())
    _bank_upload_cache[file_token] = (f.filename, file_bytes)

    return render_template("bank_confirm.html",
                           account=matched_account,
                           accounts=accounts,
                           result=result,
                           filename=f.filename,
                           file_token=file_token,
                           transactions=result.get("transactions", []))


# Cache uploaded files briefly for the confirm step (token → (filename, bytes))
_bank_upload_cache = {}


@app.route("/banks/upload/confirm", methods=["POST"])
@require_recon_auth
def bank_upload_confirm():
    account_id = request.form.get("bank_account_id", type=int)
    account = queries.get_bank_account(account_id) if account_id else None

    if not account:
        # No existing account selected — try to create from detected/edited fields
        new_bank_name = (request.form.get("new_bank_name") or "").strip()
        new_acct_num  = (request.form.get("new_account_number") or "").strip()
        new_currency  = (request.form.get("new_currency") or "USD").strip()
        new_label     = (request.form.get("new_account_label") or "").strip()
        new_entity    = (request.form.get("new_entity") or "").strip()

        if new_bank_name or new_acct_num:
            account_id = queries.create_bank_account({
                "bank_name":      new_bank_name or "Unknown",
                "account_number": new_acct_num  or "UNKNOWN",
                "account_label":  new_label,
                "currency":       new_currency,
                "entity":         new_entity,
            })
            account = queries.get_bank_account(account_id)
        else:
            return render_template("bank_upload.html",
                                   accounts=queries.get_bank_accounts(),
                                   error="No bank account selected or detected from document.")

    # Collect transactions from the form
    tx_count = int(request.form.get("tx_count", 0))
    transactions = []
    for i in range(tx_count):
        date_val = request.form.get(f"tx_date_{i}", "")
        amount_val = request.form.get(f"tx_amount_{i}", "")
        if not date_val or not amount_val:
            continue
        transactions.append({
            "date":        date_val,
            "value_date":  request.form.get(f"tx_value_date_{i}", "") or None,
            "amount":      float(amount_val),
            "balance":     float(request.form.get(f"tx_balance_{i}", "") or 0) or None,
            "currency":    request.form.get(f"tx_currency_{i}", "") or account.get("currency"),
            "reference":   request.form.get(f"tx_reference_{i}", ""),
            "description": request.form.get(f"tx_description_{i}", ""),
            "tx_type":     request.form.get(f"tx_type_{i}", "other"),
            "counterparty": request.form.get(f"tx_counterparty_{i}", ""),
        })

    if not transactions:
        return render_template("bank_upload.html",
                               accounts=queries.get_bank_accounts(),
                               error="No valid transactions to save.")

    # Compute summary
    credits = sum(t["amount"] for t in transactions if t["amount"] > 0)
    debits  = sum(t["amount"] for t in transactions if t["amount"] < 0)
    dates   = [t["date"] for t in transactions if t["date"]]

    data = {
        "bank_account_id": account_id,
        "period_start":    min(dates) if dates else None,
        "period_end":      max(dates) if dates else None,
        "filename":        request.form.get("filename", ""),
        "opening_balance": float(request.form.get("opening_balance", "") or 0) or None,
        "closing_balance": float(request.form.get("closing_balance", "") or 0) or None,
        "total_credits":   round(credits, 2),
        "total_debits":    round(debits, 2),
        "source":          request.form.get("source", "upload"),
        "transactions":    transactions,
    }

    # Retrieve cached file bytes for storage (enables download later)
    file_token = request.form.get("file_token", "")
    cached = _bank_upload_cache.pop(file_token, None)
    file_data = cached[1] if cached else None

    stmt_id = queries.create_bank_statement(data, file_data=file_data)
    return redirect(url_for("bank_detail", statement_id=stmt_id))


@app.route("/banks/add")
@require_recon_auth
def bank_manual_entry():
    accounts = queries.get_bank_accounts()
    return render_template("bank_confirm.html",
                           account=None,
                           accounts=accounts,
                           result={},
                           filename="",
                           file_token="",
                           transactions=[],
                           manual_mode=True)


@app.route("/banks/<int:statement_id>")
@require_recon_auth
def bank_detail(statement_id):
    stmt = queries.get_bank_statement(statement_id)
    if not stmt:
        abort(404)
    txns = queries.get_bank_transactions(statement_id)
    return render_template("bank_detail.html", statement=stmt, transactions=txns)


@app.route("/banks/<int:statement_id>/download")
@require_recon_auth
def bank_download(statement_id):
    filename, file_data = queries.get_bank_statement_file(statement_id)
    if not file_data:
        abort(404)
    return send_file(io.BytesIO(file_data), download_name=filename, as_attachment=True)


@app.route("/banks/<int:statement_id>/delete", methods=["POST"])
@require_recon_auth
def bank_delete(statement_id):
    queries.delete_bank_statement(statement_id)
    return redirect(url_for("banks_main"))


@app.route("/banks/transactions/<int:tx_id>/edit", methods=["POST"])
@require_recon_auth
def bank_tx_edit(tx_id):
    data = {
        "tx_date":     request.form.get("tx_date", ""),
        "value_date":  request.form.get("value_date", "") or None,
        "amount":      float(request.form.get("amount", 0)),
        "balance":     float(request.form.get("balance") or 0) or None,
        "currency":    request.form.get("currency", "").strip().upper() or None,
        "reference":   request.form.get("reference", ""),
        "description": request.form.get("description", ""),
        "tx_type":     request.form.get("tx_type", "other"),
    }
    stmt_id = queries.get_bank_tx_statement_id(tx_id)
    queries.update_bank_transaction(tx_id, data)
    if stmt_id:
        return redirect(url_for("bank_detail", statement_id=stmt_id))
    return redirect(url_for("banks_main"))


@app.route("/banks/transactions/<int:tx_id>/delete", methods=["POST"])
@require_recon_auth
def bank_tx_delete(tx_id):
    stmt_id = queries.get_bank_tx_statement_id(tx_id)
    queries.delete_bank_transaction(tx_id)
    if stmt_id:
        return redirect(url_for("bank_detail", statement_id=stmt_id))
    return redirect(url_for("banks_main"))


@app.route("/banks/<int:statement_id>/restore", methods=["POST"])
@require_recon_auth
def bank_restore(statement_id):
    queries.restore_bank_statement(statement_id)
    return redirect(url_for("banks_main"))


@app.route("/banks/<int:statement_id>/purge", methods=["POST"])
@require_recon_auth
def bank_purge(statement_id):
    queries.purge_bank_statement(statement_id)
    return redirect(url_for("banks_main"))


# ── CRO All in One Dashboard ─────────────────────────────────────────────────

@app.route("/cro")
@require_cro_auth
def cro_page():
    from flask import make_response
    resp = make_response(render_template("cro_dashboard.html"))
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate"
    return resp


@app.route("/cro/feed", methods=["POST"])
def cro_feed():
    """Unauthenticated (by session) but gated by X-Bridge-Secret header.
    Accepts live MT5 position summary from the Windows bridge pusher.
    """
    bridge_secret = os.environ.get("CRO_BRIDGE_SECRET", "")
    incoming = request.headers.get("X-Bridge-Secret", "")
    if not bridge_secret or not hmac.compare_digest(incoming, bridge_secret):
        return jsonify({"ok": False}), 403
    data = request.get_json(force=True, silent=True) or {}
    with _CRO_LIVE_LOCK:
        _CRO_LIVE.clear()
        _CRO_LIVE.update(data)
        _CRO_LIVE["received_at"] = datetime.utcnow().isoformat()
    return jsonify({"ok": True})


@app.route("/cro/refresh", methods=["POST"])
@require_cro_auth
def cro_refresh():
    """No-op now that /cro/data serves live bridge data only. Kept for
    backward-compat with the UI Refresh button -- just returns the last push."""
    with _CRO_LIVE_LOCK:
        return jsonify({"ok": True, "last_push": _CRO_LIVE.get("pushed_at")})


@app.route("/cro/data")
@require_cro_auth
def cro_data():
    """Serve the CRO dashboard data entirely from the live MT5 bridge feed.

    Dealio is no longer consulted -- the Wine-hosted cro-bridge container
    computes everything from the MT5 Manager API and POSTs it here every 30s.
    If the bridge hasn't pushed recently (>90s), we return `live_stale: true`
    so the UI can show a warning badge.
    """
    with _CRO_LIVE_LOCK:
        live = dict(_CRO_LIVE)

    if not live or not live.get("received_at"):
        return jsonify({
            "error": "MT5 bridge has not pushed yet; waiting for first cycle...",
            "live_stale": True,
        }), 503

    try:
        age = (datetime.utcnow() - datetime.fromisoformat(live["received_at"])).total_seconds()
    except Exception:
        age = 999999

    today = date.today().isoformat()
    flt = lambda k: float(live.get(k, 0) or 0)
    i   = lambda k: int(live.get(k, 0) or 0)

    daily = {
        "label":         today,
        "start":         today,
        "end":           today,
        "source":        live.get("source", "AN100"),
        "group_mask":    live.get("group_mask", "CMV*"),
        # money
        "pnl":           flt("floating_pnl_usd") + flt("closed_pnl_usd"),
        "floating_pnl":  flt("floating_pnl_usd"),
        "closed_pnl":    flt("closed_pnl_usd"),
        "delta_floating": 0.0,
        "net_deposits":  flt("net_deposits"),
        "deposits":      flt("deposits"),
        "withdrawals":   flt("withdrawals"),
        "volume_usd":    flt("volume_usd"),
        "swap":          flt("swap"),
        "commission":    flt("commission"),
        # equity snapshot -- not yet fetched live; zero for now
        "equity":        0.0,
        "balance":       0.0,
        "credit":        0.0,
        "wd_equity":     0.0,
        # counts
        "n_accounts":        i("n_traders"),
        "n_positions":       i("n_positions"),
        "n_traders":         i("n_traders"),
        "n_active_traders":  i("n_active_traders"),
        "n_depositors":      i("n_depositors"),
        "n_ftd":             0,
        "n_retention_depositors": i("n_depositors"),
        "n_deals":           i("n_closing_deals"),
    }

    # Monthly / trend / by_group are deferred until the bridge computes them.
    # For now, UI will render zeros or "No rows".
    empty_monthly = dict(daily)
    empty_monthly["label"] = today[:7]

    return jsonify({
        "date":       today,
        "requested":  today,
        "fellback":   False,
        "source":     live.get("source", "AN100"),
        "group_mask": live.get("group_mask", "CMV*"),
        "daily":      daily,
        "monthly":    empty_monthly,
        "by_group":   [],
        "by_symbol":  live.get("by_symbol", []),
        "trend":      [],
        "live_pushed_at": live.get("pushed_at"),
        "live_stale":  age >= _CRO_LIVE_MAX_AGE_S,
        "live_age_s": age,
    })


# ── Operators Dashboard ──────────────────────────────────────────────────────

@app.route("/operators")
@require_retention_auth
def operators_page():
    from flask import make_response
    resp = make_response(render_template("operators.html"))
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate"
    return resp


# ── Operator department / role derivation ─────────────────────────────────
# CRM data is inconsistent: sometimes `department` holds the functional role
# ("Conversion", "Retention") and `fax` holds a location code ("BG", "CY"),
# and sometimes it's reversed. Both fields must be checked.

_FUNC_TO_DEPT = {
    "conversion": "Sales", "retention": "Retention", "support": "Support",
    "ib": "IB", "sales pool": "Sales", "conversion bdm": "Sales",
    "retention bdm": "Retention",
}

_POS_DEPT = {
    "sales manager": "Sales", "sales vp": "Sales",
    "retention manager": "Retention", "support manager": "Support",
    "office manager": "Management", "office manager + telephones": "Management",
    "funding manager": "Operations", "compliance officer": "Compliance",
    "affiliate manager": "Affiliates", "user id telephone search": "Operations",
    "whatsapp robi": "Operations", "leaderboard": "Operations",
}

_FUNCTIONAL_ROLES = {
    "conversion", "retention", "support", "ib", "general", "sales pool",
    "white label", "white label conversion", "white label retention",
    "white label sales",
}


def _derive_department(position: str, fax: str, dept_raw: str) -> str:
    """Derive a clean department (Sales, Retention, …) from position + both fields."""
    pos = (position or "").strip().lower()
    fax_l = (fax or "").strip().lower()
    dept_l = (dept_raw or "").strip().lower()
    # 1. Position-based (managers, VPs)
    d = _POS_DEPT.get(pos)
    if d:
        return d
    # 2. Check dept field for functional keyword
    if dept_l.startswith("white label"):
        return "White Label"
    d = _FUNC_TO_DEPT.get(dept_l)
    if d:
        return d
    # 3. Check fax field for functional keyword
    if fax_l.startswith("white label"):
        return "White Label"
    d = _FUNC_TO_DEPT.get(fax_l)
    if d:
        return d
    return ""


def _is_functional(s: str) -> bool:
    low = s.strip().lower()
    return low in _FUNCTIONAL_ROLES or low.startswith("white label")


def _derive_role(fax: str, dept_raw: str) -> str:
    """Return the functional role from whichever of fax/dept contains it."""
    f = (fax or "").strip()
    d = (dept_raw or "").strip()
    if _is_functional(f):
        return f
    if _is_functional(d):
        return d
    return ""


_ROLE_NORMALIZE = {
    "ret": "Retention", "ret ": "Retention", "re": "Retention",
    "retention ": "Retention", "conv": "Conversion", "con": "Conversion",
}
_ROLE_SUPPRESS = {"zzz", "null", "junk", "closed", "il (closed)", "bg (closed)",
                  "test qa", "qa desk 2", ""}


def _clean_role(raw: str) -> str:
    """Normalize abbreviations, suppress junk values."""
    if not raw:
        return ""
    s = raw.strip()
    low = s.lower()
    if low in _ROLE_SUPPRESS:
        return ""
    return _ROLE_NORMALIZE.get(low, s)


@app.route("/operators/data")
@require_retention_auth
def operators_data():
    """Fast endpoint — just operator list from vtiger_users (<2s)."""
    try:
        ops = queries.operator_list()
        for op in ops:
            raw_dept = op.get("department") or ""
            raw_fax = op.get("role_name") or ""   # SQL: fax AS role_name
            op["team"] = raw_dept                  # preserve original CRM team code
            op["department"] = _derive_department(op.get("position"), raw_fax, raw_dept)
            op["role_name"] = _clean_role(_derive_role(raw_fax, raw_dept))
            op["last_login"] = str(op.get("last_login") or "")[:16]
            if op.get("position") == "Agent":
                op["position"] = "Sales Agent"
        return jsonify({"operators": ops})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e), "operators": []})


@app.route("/operators/watermark")
@require_retention_auth
def operators_watermark():
    """Query the max operator ID watermark for new-operator detection.

    The watermark is normally updated by the daily cron job
    (snapshot_watermark.py at 00:01). This endpoint only seeds an initial
    baseline when the table is empty (prev_max == 0 — first ever run before
    the cron has fired). After that it is read-only so the cron's morning
    snapshot is not overwritten by mid-day page loads.
    """
    import datetime as _dt
    current_max = request.args.get("current_max", type=int, default=0)
    stored = queries.get_operator_watermark()
    prev_max  = stored["max_id"]
    prev_date = (stored.get("updated_at") or "")[:10]

    # First-ever run: seed the baseline so the dashboard isn't blank until midnight
    if prev_max == 0 and current_max > 0:
        queries.update_operator_watermark(current_max)
        return jsonify({
            "prev_max_id":     current_max,
            "prev_date":       _dt.date.today().isoformat(),
            "current_max_id":  current_max,
            "new_count":       0,
            "explanation":     "Baseline set: ID #%d. Cron will snapshot at midnight — "
                               "new operators will be tracked from tomorrow." % current_max,
        })

    new_count = max(0, current_max - prev_max)
    explanation = (
        "Operators with ID > #%d (snapshot: %s) are new — %d found"
        % (prev_max, prev_date, new_count)
    )
    return jsonify({
        "prev_max_id":     prev_max,
        "prev_date":       prev_date,
        "current_max_id":  current_max,
        "new_count":       new_count,
        "explanation":     explanation,
    })


@app.route("/operators/stats")
@require_retention_auth
def operators_stats():
    """Slow endpoint — client + FTD stats from vtiger_account (can take 30-60s)."""
    import concurrent.futures as _cf
    try:
        with _cf.ThreadPoolExecutor(max_workers=2) as ex:
            f_cs  = ex.submit(queries.operator_client_stats)
            f_ftd = ex.submit(queries.operator_ftd_stats)
            client_stats = f_cs.result()
            ftd_stats    = f_ftd.result()

        merged = {}
        for oid in set(list(client_stats.keys()) + list(ftd_stats.keys())):
            cs = client_stats.get(oid, {})
            fs = ftd_stats.get(oid, {})
            merged[str(oid)] = {
                "total_clients":        int(cs.get("total_clients") or 0),
                "funded_clients":       int(cs.get("funded_clients") or 0),
                "total_deposit_volume": round(float(cs.get("total_deposit_volume") or 0), 2),
                "net_deposit_volume":   round(float(cs.get("net_deposit_volume") or 0), 2),
                "ftd_count":            int(fs.get("ftd_count") or 0),
                "ftd_volume":           round(float(fs.get("ftd_volume") or 0), 2),
            }
        return jsonify({"stats": merged})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e), "stats": {}})


if __name__ == "__main__":
    debug = os.environ.get("FLASK_DEBUG", "0") == "1"
    app.run(host="0.0.0.0", port=5050, debug=debug)
