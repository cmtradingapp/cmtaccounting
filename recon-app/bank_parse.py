"""
Bank statement parsers — CSV, XLS, XLSX, and PDF (AI vision).

Each structured format (CSV/XLS/XLSX) has per-bank detection + extraction.
PDFs fall back to Claude vision (same infrastructure as ai_parse.py).
"""

import csv
import io
import json
import re
from datetime import datetime


# ═══════════════════════════════════════════════════════════════════════════════
# FORMAT DETECTION
# ═══════════════════════════════════════════════════════════════════════════════

def detect_bank_format(file_bytes, filename):
    """Auto-detect bank and format from file content / filename.
    Returns a string like 'nedbank_csv', 'absa_xls', 'generic_csv', etc.
    """
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    name_lower = filename.lower()

    if ext == "csv":
        text = file_bytes.decode("utf-8", errors="replace")
        if "1053484658" in text or "1127489763" in text or "nedbank" in text.lower():
            return "nedbank_csv"
        if "280544308" in text or "standard bank" in text.lower():
            return "standard_csv"
        return "generic_csv"

    if ext in ("xls", "xlsx"):
        # Try to peek at content for bank identification
        text_preview = file_bytes[:4096].decode("utf-8", errors="replace").lower()
        if "absa" in text_preview or "absa" in name_lower:
            return "absa_xls"
        if "gtbank" in text_preview or "gtbank" in name_lower or "guaranty" in text_preview:
            return "gtbank_xls"
        if "access" in text_preview or "access bank" in name_lower:
            return "access_xls"
        if "adib" in text_preview or "abu dhabi" in text_preview or "adib" in name_lower:
            return "adib_xls"
        if "dixipay" in text_preview or "dixipay" in name_lower:
            return "dixipay_xlsx"
        if "nuvei" in text_preview or "neteller" in text_preview or "nuvei" in name_lower:
            return "nuvei_xlsx"
        return "generic_xlsx"

    if ext == "pdf":
        return "pdf"

    return "unknown"


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════════

_BANK_META_SYSTEM_PROMPT = """\
You are a bank statement header parser for a financial reconciliation system.

Given the first section of a bank statement (CSV, XLS, or HTML-exported file),
extract ONLY the account metadata fields listed below. The transactions have already
been extracted by a deterministic parser — do NOT re-parse transactions.

Return ONLY a single valid JSON object. No markdown fences, no explanation, no extra keys.

═══════════════════════════════════════════════════════════
  OUTPUT SCHEMA  (use null for any unknown field)
═══════════════════════════════════════════════════════════
{
  "bank_name":       "<short readable bank name, e.g. 'Guaranty Trust Bank' or 'Nedbank'>",
  "account_number":  "<digits only, keep leading zeros, e.g. '0624057342' or '280544308'>",
  "entity_name":     "<account holder name, e.g. 'CM FINTECH NIGERIA LTD' or 'GCMT SA PTY'>",
  "currency":        "<ISO-4217 code, e.g. 'NGN', 'ZAR', 'USD', 'AED', 'EUR', 'GBP', 'GHS', 'KES'>",
  "opening_balance": <number or null>,
  "closing_balance": <number or null>
}

═══════════════════════════════════════════════════════════
  RULES
═══════════════════════════════════════════════════════════
- bank_name: the institution's name as printed, not a product name
- account_number: digits only — strip spaces, dashes, dots; KEEP leading zeros
- entity_name: the account holder / company name
- currency: infer from country context if not stated explicitly
  (Nigeria → NGN, South Africa → ZAR, UAE → AED, Ghana → GHS, Kenya → KES)
- opening_balance / closing_balance: the stated balance figures, as plain numbers
- If a field is genuinely absent, use null — never guess randomly
- Return ONLY the JSON object
"""


def _ai_enrich_metadata(file_bytes: bytes, filename: str, result: dict) -> None:
    """Call Claude to fill in any metadata fields the heuristic scan missed.
    Modifies `result` in-place, only overwriting empty/null/Unknown values.
    Called only when at least one key field is missing.
    """
    try:
        import ai_parse as _ai
    except ImportError:
        return

    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    # Build a text sample from the file header (not transactions — already parsed)
    try:
        if ext in ("xls",) or b"<html" in file_bytes[:512].lower() or b"<!doc" in file_bytes[:512].lower():
            # HTML-as-XLS: strip tags
            raw = file_bytes.decode("utf-8", errors="replace")
            sample = re.sub(r"<[^>]+>", " ", raw)
            sample = re.sub(r"&nbsp;|&amp;|&lt;|&gt;", " ", sample)
            sample = re.sub(r"\s+", " ", sample)[:3000]
        elif ext == "xlsx":
            # XLSX: render first 25 rows as tab-separated
            rows = _read_xlsx(file_bytes)
            lines = ["\t".join(str(c or "") for c in row) for row in rows[:25]]
            sample = "\n".join(lines)
        else:
            # CSV / plain text
            sample = file_bytes.decode("utf-8", errors="replace")[:3000]
    except Exception:
        return

    user_msg = (
        f"File: {filename}\n\n"
        f"Header section of the bank statement:\n\n{sample}\n\n"
        f"Extract the account metadata fields as JSON."
    )

    try:
        raw = _ai._call_claude(_BANK_META_SYSTEM_PROMPT, user_msg)
        import json as _json
        # Handle markdown fences if present
        cleaned = re.sub(r"^```\w*\s*", "", raw.strip())
        cleaned = re.sub(r"\s*```$", "", cleaned)
        data = _json.loads(cleaned)
    except Exception:
        return

    # Merge: only fill fields that are empty / null / "Unknown"
    _EMPTY = {"", "Unknown", "unknown", None}
    for field in ("bank_name", "account_number", "entity_name", "currency"):
        if result.get(field) in _EMPTY and data.get(field):
            result[field] = str(data[field]).strip()
    for field in ("opening_balance", "closing_balance"):
        if result.get(field) is None and data.get(field) is not None:
            try:
                result[field] = float(data[field])
            except (TypeError, ValueError):
                pass


def parse_bank_statement(file_bytes, filename, bank_format=None, vision_mode=False):
    """Parse a bank statement file and return structured data.

    Returns: {
        "bank_name": str,
        "account_number": str,
        "currency": str,
        "opening_balance": float or None,
        "closing_balance": float or None,
        "transactions": [
            {"date": "YYYY-MM-DD", "amount": float, "balance": float,
             "reference": str, "description": str, "tx_type": str, ...},
        ]
    }
    """
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if not bank_format:
        bank_format = detect_bank_format(file_bytes, filename)

    # PDF → AI vision
    if ext == "pdf" or bank_format == "pdf":
        if vision_mode:
            return _parse_pdf_vision(file_bytes, filename)
        return _parse_pdf_text(file_bytes, filename)

    # CSV
    if ext == "csv":
        if bank_format == "nedbank_csv":
            result = _parse_nedbank_csv(file_bytes)
        elif bank_format == "standard_csv":
            result = _parse_standard_csv(file_bytes)
        else:
            result = _parse_generic_csv(file_bytes)
    # XLS / XLSX
    elif ext in ("xls", "xlsx"):
        result = _parse_excel(file_bytes, filename, bank_format)
    else:
        raise ValueError(f"Unsupported file format: {ext}")

    # AI enrichment: fill in any metadata fields the heuristic missed
    _needs_enrichment = (
        result.get("bank_name", "Unknown") in ("", "Unknown") or
        not result.get("account_number") or
        not result.get("currency")
    )
    if _needs_enrichment:
        _ai_enrich_metadata(file_bytes, filename, result)

    return result


# ═══════════════════════════════════════════════════════════════════════════════
# CSV PARSERS
# ═══════════════════════════════════════════════════════════════════════════════

def _clean_amount(val):
    """Parse a monetary value string → float. Handles commas, spaces, parens for negative."""
    if not val or not str(val).strip():
        return None
    s = str(val).strip()
    # Parentheses = negative: (1,234.56) → -1234.56
    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg = True
        s = s[1:-1]
    s = re.sub(r"[^\d.\-]", "", s)
    try:
        v = float(s)
        return -v if neg else v
    except ValueError:
        return None


def _parse_date(val):
    """Try common date formats and return YYYY-MM-DD."""
    if not val:
        return None
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d",
                "%d-%m-%Y", "%d-%b-%Y", "%d-%B-%Y",
                "%d %b %Y", "%d %B %Y", "%Y%m%d",
                "%b %d, %Y", "%B %d, %Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s  # return as-is if no format matched


def _guess_tx_type(amount, desc=""):
    """Infer transaction type from amount sign and description keywords."""
    desc_lower = (desc or "").lower()
    if "fee" in desc_lower or "charge" in desc_lower or "commission" in desc_lower:
        return "fee"
    if "interest" in desc_lower:
        return "interest"
    if "transfer" in desc_lower:
        return "transfer"
    if amount is not None:
        return "deposit" if amount > 0 else "withdrawal"
    return "other"


def _parse_nedbank_csv(file_bytes):
    """Parse Nedbank CSV statement (ZAR accounts)."""
    text = file_bytes.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)

    result = {
        "bank_name": "Nedbank",
        "account_number": "",
        "entity_name": "",
        "currency": "ZAR",
        "opening_balance": None,
        "closing_balance": None,
        "transactions": [],
    }

    # Find header row
    header_idx = None
    for i, row in enumerate(rows):
        row_lower = [str(c).lower().strip() for c in row]
        if "date" in row_lower and ("amount" in row_lower or "balance" in row_lower):
            header_idx = i
            break

    if header_idx is None:
        # Fallback: try to detect columns by position
        return _parse_generic_csv(file_bytes)

    headers = [str(c).strip().lower() for c in rows[header_idx]]

    # Map columns
    col_map = {}
    for i, h in enumerate(headers):
        if "date" in h and "value" not in h:
            col_map.setdefault("date", i)
        elif "value" in h and "date" in h:
            col_map["value_date"] = i
        elif h in ("amount", "debit/credit", "transaction amount"):
            col_map["amount"] = i
        elif "balance" in h:
            col_map["balance"] = i
        elif "description" in h or "narrative" in h or "detail" in h:
            col_map.setdefault("description", i)
        elif "reference" in h or "ref" in h:
            col_map.setdefault("reference", i)

    # Extract account number from early rows
    for row in rows[:header_idx]:
        for cell in row:
            cell_s = str(cell).strip()
            if re.match(r"^\d{8,12}$", cell_s):
                result["account_number"] = cell_s
                break

    # Parse data rows
    for row in rows[header_idx + 1:]:
        if len(row) < 2:
            continue
        date_val = _parse_date(row[col_map["date"]]) if "date" in col_map else None
        if not date_val:
            continue
        amount = _clean_amount(row[col_map["amount"]]) if "amount" in col_map else None
        if amount is None:
            continue

        tx = {
            "date": date_val,
            "value_date": _parse_date(row[col_map.get("value_date", -1)]) if "value_date" in col_map else None,
            "amount": amount,
            "balance": _clean_amount(row[col_map["balance"]]) if "balance" in col_map else None,
            "reference": str(row[col_map["reference"]]).strip() if "reference" in col_map and col_map["reference"] < len(row) else "",
            "description": str(row[col_map["description"]]).strip() if "description" in col_map and col_map["description"] < len(row) else "",
            "tx_type": _guess_tx_type(amount, str(row[col_map.get("description", 0)]) if "description" in col_map else ""),
        }
        result["transactions"].append(tx)

    # Compute opening/closing from balance column
    if result["transactions"]:
        balances = [t["balance"] for t in result["transactions"] if t.get("balance") is not None]
        if balances:
            first_tx = result["transactions"][0]
            if first_tx.get("balance") is not None and first_tx.get("amount") is not None:
                result["opening_balance"] = round(first_tx["balance"] - first_tx["amount"], 2)
            result["closing_balance"] = balances[-1]

    return result


def _parse_standard_csv(file_bytes):
    """Parse Standard Bank CSV statement."""
    result = _parse_generic_csv(file_bytes)
    result["bank_name"] = "Standard Bank"
    result.setdefault("entity_name", "")
    result.setdefault("currency", "ZAR")

    text = file_bytes.decode("utf-8", errors="replace")
    header = text[:1000]
    # Extract account number
    m = re.search(r"Account\s+0*(\d{6,12})", header, re.IGNORECASE)
    if m:
        result["account_number"] = m.group(1)
    elif re.search(r"280544308|(\d{9,12})", header):
        result["account_number"] = re.search(r"280544308|(\d{9,12})", header).group(0).lstrip("0")
    # Extract entity name (line after "Account XXXXXXXXX")
    m_entity = re.search(r"Account\s+\d+\s+([A-Z][A-Z &()./]{4,})", header)
    if m_entity:
        result["entity_name"] = m_entity.group(1).strip()
    return result


def _parse_generic_csv(file_bytes):
    """Generic CSV parser — tries to detect columns intelligently."""
    text = file_bytes.decode("utf-8", errors="replace")

    # Try different delimiters
    for delimiter in [",", ";", "\t", "|"]:
        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        rows = list(reader)
        if rows and max(len(r) for r in rows) >= 3:
            break

    result = {
        "bank_name": "Unknown",
        "account_number": "",
        "entity_name": "",
        "currency": "",
        "opening_balance": None,
        "closing_balance": None,
        "transactions": [],
    }

    if not rows:
        return result

    # Find the header row (first row with at least 3 non-empty cells that looks like headers)
    header_idx = 0
    for i, row in enumerate(rows[:10]):
        non_empty = [c for c in row if str(c).strip()]
        row_text = " ".join(str(c).lower() for c in row)
        if len(non_empty) >= 3 and ("date" in row_text or "amount" in row_text):
            header_idx = i
            break

    headers = [str(c).strip().lower() for c in rows[header_idx]]

    # Auto-detect column mapping
    col_map = {}
    for i, h in enumerate(headers):
        if not col_map.get("date") and ("date" in h and "value" not in h):
            col_map["date"] = i
        elif "value" in h and "date" in h:
            col_map["value_date"] = i
        elif not col_map.get("amount") and h in ("amount", "debit/credit", "sum", "transaction amount"):
            col_map["amount"] = i
        elif not col_map.get("debit") and ("debit" in h or "dr" == h):
            col_map["debit"] = i
        elif not col_map.get("credit") and ("credit" in h or "cr" == h):
            col_map["credit"] = i
        elif not col_map.get("balance") and "balance" in h:
            col_map["balance"] = i
        elif not col_map.get("description") and any(k in h for k in ("desc", "narrative", "detail", "particular", "remark", "narr", "memo", "note", "transaction detail")):
            col_map["description"] = i
        elif not col_map.get("reference") and ("ref" in h or "reference" in h):
            col_map["reference"] = i

    # If no "amount" column but we have debit/credit, combine them
    has_split = "debit" in col_map and "credit" in col_map
    if "amount" not in col_map and not has_split:
        # Try to find a numeric column
        for i, h in enumerate(headers):
            if i not in col_map.values():
                col_map["amount"] = i
                break

    for row in rows[header_idx + 1:]:
        if len(row) < 2 or not any(str(c).strip() for c in row):
            continue

        date_val = _parse_date(row[col_map["date"]]) if "date" in col_map and col_map["date"] < len(row) else None
        if not date_val:
            continue

        if has_split:
            debit = _clean_amount(row[col_map["debit"]]) if col_map["debit"] < len(row) else None
            credit = _clean_amount(row[col_map["credit"]]) if col_map["credit"] < len(row) else None
            amount = (credit or 0) - (debit or 0) if (credit or debit) else None
        else:
            amount = _clean_amount(row[col_map["amount"]]) if "amount" in col_map and col_map["amount"] < len(row) else None

        if amount is None:
            continue

        desc = str(row[col_map["description"]]).strip() if "description" in col_map and col_map["description"] < len(row) else ""

        tx = {
            "date": date_val,
            "value_date": _parse_date(row[col_map.get("value_date", -1)]) if "value_date" in col_map and col_map.get("value_date", 999) < len(row) else None,
            "amount": amount,
            "balance": _clean_amount(row[col_map["balance"]]) if "balance" in col_map and col_map["balance"] < len(row) else None,
            "reference": str(row[col_map["reference"]]).strip() if "reference" in col_map and col_map["reference"] < len(row) else "",
            "description": desc,
            "tx_type": _guess_tx_type(amount, desc),
        }
        result["transactions"].append(tx)

    if result["transactions"]:
        balances = [t["balance"] for t in result["transactions"] if t.get("balance") is not None]
        if balances:
            first_tx = result["transactions"][0]
            if first_tx.get("balance") is not None and first_tx.get("amount") is not None:
                result["opening_balance"] = round(first_tx["balance"] - first_tx["amount"], 2)
            result["closing_balance"] = balances[-1]

    return result


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL PARSERS (XLS / XLSX)
# ═══════════════════════════════════════════════════════════════════════════════

def _extract_html_metadata(file_bytes):
    """Extract bank metadata from HTML-as-XLS files using regex on plain text."""
    text = file_bytes.decode("utf-8", errors="replace")
    # Strip HTML tags to get plain text
    plain = re.sub(r"<[^>]+>", " ", text)
    plain = re.sub(r"&nbsp;", " ", plain)
    plain = re.sub(r"&[a-z]+;", " ", plain)
    plain = re.sub(r"\s+", " ", plain)

    meta = {}

    # Bank name — stop at CUSTOMER/STATEMENT/PLC suffixes
    m = re.search(r"([A-Z][A-Z &]+(?:BANK|FINANCE|TRUST)(?:\s+(?:PLC|LTD|INC|CORP|GROUP))?)", plain)
    if m:
        meta["bank_name"] = m.group(1).strip().title()

    # Account number — keep leading zeros (significant in e.g. Nigerian account numbers)
    m = re.search(r"Account\s+No[.:\s]+([0-9A-Z\-/]{6,20})", plain, re.IGNORECASE)
    if m:
        acct = re.sub(r"[^0-9A-Z]", "", m.group(1).upper())
        meta["account_number"] = acct

    # Account / entity name
    m = re.search(r"Account\s+Name[:\s]+([A-Z][A-Za-z &().\-]{3,60}?)(?:\s{2,}|Print|Address|Account\s+No)", plain, re.IGNORECASE)
    if m:
        meta["entity_name"] = m.group(1).strip()

    # Currency — common name → ISO code
    _CCY_MAP = {
        "naira": "NGN", "dollar": "USD", "euro": "EUR", "pound": "GBP",
        "rand": "ZAR", "dirham": "AED", "cedi": "GHS", "shilling": "KES",
        "franc": "XOF",
    }
    m = re.search(r"Currency[:\s]+([A-Za-z\s]{3,20}?)(?:\s{2,}|\n|Period|Opening)", plain, re.IGNORECASE)
    if m:
        ccy_raw = m.group(1).strip().lower()
        meta["currency"] = _CCY_MAP.get(ccy_raw, ccy_raw.upper()[:3])
    # Also check for explicit ISO code
    m2 = re.search(r"\b(NGN|USD|EUR|GBP|ZAR|AED|GHS|KES|XOF|XAF)\b", plain)
    if m2 and "currency" not in meta:
        meta["currency"] = m2.group(1)

    # Opening balance
    m = re.search(r"Opening\s+Balance[:\s]+([\d,]+\.?\d*)", plain, re.IGNORECASE)
    if m:
        try:
            meta["opening_balance"] = float(m.group(1).replace(",", ""))
        except ValueError:
            pass

    return meta


def _parse_excel(file_bytes, filename, bank_format):
    """Parse XLS/XLSX using openpyxl (xlsx) or xlrd (xls)."""
    ext = filename.rsplit(".", 1)[-1].lower()

    if ext == "xlsx":
        rows = _read_xlsx(file_bytes)
    elif ext == "xls":
        rows = _read_xls(file_bytes)
    else:
        raise ValueError(f"Unsupported Excel format: {ext}")

    # For HTML-as-XLS files, extract header metadata before parsing the table
    html_meta = {}
    preview = file_bytes[:512].lstrip()
    if preview.startswith(b"<") or b"<!DOCTYPE" in preview or b"<html" in preview[:512].lower():
        html_meta = _extract_html_metadata(file_bytes)

    # Convert to the same structure as CSV and parse generically
    _bank_label = bank_format.replace("_xls", "").replace("_xlsx", "").replace("generic", "Unknown").title()
    result = {
        "bank_name": html_meta.get("bank_name") or _bank_label,
        "account_number": html_meta.get("account_number", ""),
        "entity_name": html_meta.get("entity_name", ""),
        "currency": html_meta.get("currency", ""),
        "opening_balance": html_meta.get("opening_balance"),
        "closing_balance": None,
        "transactions": [],
    }

    if not rows:
        return result

    # Find header row
    header_idx = 0
    for i, row in enumerate(rows[:15]):
        row_text = " ".join(str(c).lower() for c in row if c is not None)
        non_empty = [c for c in row if c is not None and str(c).strip()]
        if len(non_empty) >= 3 and ("date" in row_text or "amount" in row_text or "balance" in row_text):
            header_idx = i
            break

    headers = [str(c).strip().lower() if c else "" for c in rows[header_idx]]

    # Auto-detect columns (same logic as generic CSV)
    col_map = {}
    for i, h in enumerate(headers):
        if not col_map.get("date") and ("date" in h and "value" not in h):
            col_map["date"] = i
        elif "value" in h and "date" in h:
            col_map["value_date"] = i
        elif not col_map.get("amount") and h in ("amount", "debit/credit", "sum", "transaction amount"):
            col_map["amount"] = i
        elif not col_map.get("debit") and ("debit" in h or "dr" == h):
            col_map["debit"] = i
        elif not col_map.get("credit") and ("credit" in h or "cr" == h):
            col_map["credit"] = i
        elif not col_map.get("balance") and "balance" in h:
            col_map["balance"] = i
        elif not col_map.get("description") and any(k in h for k in ("desc", "narrative", "detail", "particular", "remark", "narr", "memo", "note", "transaction detail")):
            col_map["description"] = i
        elif not col_map.get("reference") and ("ref" in h or "reference" in h):
            col_map["reference"] = i

    has_split = "debit" in col_map and "credit" in col_map

    for row in rows[header_idx + 1:]:
        if not row or not any(c is not None and str(c).strip() for c in row):
            continue

        date_val = None
        if "date" in col_map and col_map["date"] < len(row):
            cell = row[col_map["date"]]
            if isinstance(cell, datetime):
                date_val = cell.strftime("%Y-%m-%d")
            else:
                date_val = _parse_date(cell)
        if not date_val:
            continue

        if has_split:
            debit = _clean_amount(row[col_map["debit"]]) if col_map["debit"] < len(row) else None
            credit = _clean_amount(row[col_map["credit"]]) if col_map["credit"] < len(row) else None
            amount = (credit or 0) - (debit or 0) if (credit or debit) else None
        else:
            amount = _clean_amount(row[col_map["amount"]]) if "amount" in col_map and col_map["amount"] < len(row) else None

        if amount is None:
            continue

        desc = str(row[col_map["description"]]).strip() if "description" in col_map and col_map["description"] < len(row) and row[col_map["description"]] else ""

        tx = {
            "date": date_val,
            "value_date": None,
            "amount": amount,
            "balance": _clean_amount(row[col_map["balance"]]) if "balance" in col_map and col_map["balance"] < len(row) else None,
            "reference": str(row[col_map["reference"]]).strip() if "reference" in col_map and col_map["reference"] < len(row) and row[col_map["reference"]] else "",
            "description": desc,
            "tx_type": _guess_tx_type(amount, desc),
        }
        result["transactions"].append(tx)

    if result["transactions"]:
        balances = [t["balance"] for t in result["transactions"] if t.get("balance") is not None]
        if balances:
            first_tx = result["transactions"][0]
            if first_tx.get("balance") is not None and first_tx.get("amount") is not None:
                result["opening_balance"] = round(first_tx["balance"] - first_tx["amount"], 2)
            result["closing_balance"] = balances[-1]

    return result


def _read_xlsx(file_bytes):
    """Read XLSX into list of lists using openpyxl."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    wb.close()
    return rows


def _read_xls(file_bytes):
    """Read old XLS format using xlrd, with HTML-as-XLS fallback.

    Many banks (GTBank, Access Bank, etc.) export HTML tables with a .xls
    extension. Excel opens them fine; xlrd rejects them. Detect and handle.
    """
    # Detect HTML-disguised-as-XLS
    preview = file_bytes[:512].lstrip()
    if preview.startswith(b'<') or preview.startswith(b'\r\n<') or preview.startswith(b'\n<'):
        return _read_html_table(file_bytes)

    try:
        import xlrd
        wb = xlrd.open_workbook(file_contents=file_bytes)
        ws = wb.sheet_by_index(0)
        rows = []
        for i in range(ws.nrows):
            row = []
            for j in range(ws.ncols):
                cell = ws.cell(i, j)
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        dt = xlrd.xldate_as_datetime(cell.value, wb.datemode)
                        row.append(dt)
                    except Exception:
                        row.append(cell.value)
                else:
                    row.append(cell.value)
            rows.append(row)
        return rows
    except ImportError:
        raise ImportError(
            "xlrd is required for .xls files. Install with: pip install xlrd"
        )
    except Exception as e:
        # Final fallback: try HTML if xlrd chokes (e.g. "Expected BOF record")
        if b'<' in file_bytes[:1024]:
            return _read_html_table(file_bytes)
        raise


def _read_html_table(file_bytes):
    """Parse an HTML file (often exported as .xls by banks) into a list of rows."""
    import html
    from html.parser import HTMLParser

    class _TableParser(HTMLParser):
        def __init__(self):
            super().__init__()
            self.rows = []
            self._cur_row = None
            self._cur_cell = None
            self._in_cell = False

        def handle_starttag(self, tag, attrs):
            if tag == 'tr':
                self._cur_row = []
            elif tag in ('td', 'th') and self._cur_row is not None:
                self._cur_cell = []
                self._in_cell = True

        def handle_endtag(self, tag):
            if tag == 'tr' and self._cur_row is not None:
                if any(c for c in self._cur_row):
                    self.rows.append(self._cur_row)
                self._cur_row = None
            elif tag in ('td', 'th') and self._cur_row is not None:
                text = html.unescape(''.join(self._cur_cell)).strip()
                self._cur_row.append(text)
                self._cur_cell = None
                self._in_cell = False

        def handle_data(self, data):
            if self._in_cell and self._cur_cell is not None:
                self._cur_cell.append(data)

    text = file_bytes.decode('utf-8', errors='replace')
    parser = _TableParser()
    parser.feed(text)
    return parser.rows


# ═══════════════════════════════════════════════════════════════════════════════
# PDF PARSERS
# ═══════════════════════════════════════════════════════════════════════════════

_BANK_STATEMENT_SYSTEM_PROMPT = """You are a bank statement parser. Extract ALL transactions and account details from this bank statement.

Return ONLY a JSON object with this exact structure (no markdown, no explanation):
{
    "bank_name": "the bank name (e.g. Standard Bank, Nedbank, ABSA)",
    "account_number": "the account number digits only, no leading zeros",
    "entity_name": "the account holder name (e.g. GCMT SA PTY CLIENT D)",
    "currency": "the currency code (USD, ZAR, EUR, AED, NGN, etc.)",
    "opening_balance": 1234.56,
    "closing_balance": 5678.90,
    "transactions": [
        {
            "date": "YYYY-MM-DD",
            "amount": 100.00,
            "balance": 1334.56,
            "reference": "transaction reference if any",
            "description": "full transaction description",
            "tx_type": "deposit"
        }
    ]
}

Rules:
- Deposits/credits are POSITIVE amounts
- Withdrawals/debits are NEGATIVE amounts
- tx_type must be one of: deposit, withdrawal, fee, interest, transfer, other
- Dates must be YYYY-MM-DD format
- Include ALL transactions, do not skip any
- Strip leading zeros from account_number (280544308 not 0000280544308)
- If a field is not available, use null
- Return ONLY the JSON, no markdown fences, no explanation
"""


def _parse_pdf_text(file_bytes, filename):
    """Try to extract from a text-based PDF using pdfplumber."""
    import pdfplumber
    text = ""
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            text += (page.extract_text() or "") + "\n"

    if not text.strip():
        raise ValueError(
            "No text could be extracted from this PDF. "
            "Try enabling 'Scanned PDF (vision)' mode."
        )

    # Use AI to parse the extracted text
    return _parse_with_ai(text, is_pdf_bytes=None)


def _parse_pdf_vision(file_bytes, filename):
    """Use Claude vision to parse a scanned/image PDF."""
    import ai_parse
    raw = ai_parse._call_claude_with_pdf(
        file_bytes,
        _BANK_STATEMENT_SYSTEM_PROMPT,
        "Extract all transactions from this bank statement. Return JSON only."
    )
    return _parse_ai_response(raw)


def _parse_with_ai(text, is_pdf_bytes=None):
    """Send extracted text to Claude for structured extraction."""
    import ai_parse
    raw = ai_parse._call_claude(
        _BANK_STATEMENT_SYSTEM_PROMPT,
        f"Here is the bank statement text:\n\n{text}\n\nExtract all transactions as JSON."
    )
    return _parse_ai_response(raw)


def _parse_ai_response(raw):
    """Parse Claude's JSON response into our standard format."""
    # Strip markdown fences if present
    cleaned = raw.strip()
    if cleaned.startswith("```"):
        cleaned = re.sub(r"^```\w*\n?", "", cleaned)
        cleaned = re.sub(r"\n?```$", "", cleaned)

    try:
        data = json.loads(cleaned)
    except json.JSONDecodeError:
        # Try to find JSON in the response
        m = re.search(r"\{[\s\S]*\}", cleaned)
        if m:
            data = json.loads(m.group(0))
        else:
            raise ValueError("Could not parse AI response as JSON")

    # Normalize
    acct_num = str(data.get("account_number", "") or "").strip().lstrip("0") or ""
    result = {
        "bank_name": data.get("bank_name", "Unknown"),
        "account_number": acct_num,
        "entity_name": str(data.get("entity_name", "") or "").strip(),
        "currency": data.get("currency", ""),
        "opening_balance": data.get("opening_balance"),
        "closing_balance": data.get("closing_balance"),
        "transactions": [],
    }

    for tx in data.get("transactions", []):
        amount = tx.get("amount")
        if amount is None:
            continue
        result["transactions"].append({
            "date": tx.get("date", ""),
            "value_date": tx.get("value_date"),
            "amount": float(amount),
            "balance": float(tx["balance"]) if tx.get("balance") is not None else None,
            "reference": str(tx.get("reference", "") or ""),
            "description": str(tx.get("description", "") or ""),
            "tx_type": tx.get("tx_type", "other"),
        })

    return result
